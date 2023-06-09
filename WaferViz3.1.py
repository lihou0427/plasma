# Started learning python 10/23/2021 

import tkinter as tk
from tkinter import ttk
from tkinter import *
from tkinter.filedialog import askopenfilename
from PIL import ImageTk, Image

from io import BytesIO
import win32clipboard
import csv
import openpyxl

import matplotlib
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
pd.options.mode.chained_assignment = None  

from matplotlib import gridspec
matplotlib.use('TkAgg')
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from mpl_toolkits.mplot3d import axes3d
from matplotlib import cm
from scipy import stats

import math
import seaborn as sns
import os
import threading

from scipy.interpolate import griddata
from matplotlib.patches import Circle
from matplotlib.figure import Figure

plt.rcParams['font.family'] ='Arial'
plt.rcParams['font.size'] =5
plt.rcParams['axes.linewidth'] =0.2
plt.rcParams['figure.dpi'] =250
    
class app_gui:
    def __init__(self):
        self.window_main = tk.Tk()
        self.window_main.title("WaferViz 3.1 - Wafer Plotter")
        self.window_main.geometry("1484x840+130+45")
        self.window_main.resizable(1, 1)
        self.window_main.minsize(965, 640)
        self.window_main.columnconfigure(0, weight=1)
        self.window_main.rowconfigure(1, weight=1)

        self.frame_control = tk.Frame(self.window_main, borderwidth=1, relief=SUNKEN)
        self.frame_control.grid(row=0, column=0, sticky="nesw", padx=1, pady=0)
        self.frame_control.columnconfigure(0, weight=0)
        self.frame_control.rowconfigure(0, weight=0)

        self.frame_display = tk.Frame(self.window_main, borderwidth=1, relief=SUNKEN)
        self.frame_display.grid(row=1, column=0, sticky="nesw", padx=1, pady=0)
        self.frame_display.columnconfigure(0, weight=1)
        self.frame_display.rowconfigure(0, weight=1)
        
        self.button_atlas = tk.Button(self.frame_control, text="Load Atlas", command=self.open_atlas,
                            width=13)     
        self.button_atlas.grid(row=0, column=0, sticky="w", columnspan=2, padx=10, pady=0)
        self.button_workbook = tk.Button(self.frame_control, text="Load Workbook",
                            command=self.open_workbook, width=13)     
        self.button_workbook.grid(row=2, column=0, sticky="w", columnspan=2, padx=10, pady=0)
        self.file_path = tk.Label(self.frame_control, text="", justify='left', width=47, anchor="w",
                            highlightthickness=0, highlightbackground="lightgrey")
        self.file_path.grid(row=1, column=0, columnspan=6, sticky="w", padx=6, pady=0)

        self.label_id = tk.Label(self.frame_control, text="ID")
        self.label_id.grid(row=2, column=2, sticky='e', padx=4, pady=0)
        self.entry_id = tk.Entry(self.frame_control, width=12, justify='center')
        self.entry_id.insert(0, 'Wafer #')
        self.entry_id.grid(row=2, column=3, columnspan=3, sticky='w', padx=0, pady=0)

        self.radio_sigma = tk.IntVar()
        self.radio_sigma.set(1)
        self.button_sigma = tk.Radiobutton(self.frame_control, text='Data Filter \u00B1\u03C3',
                            variable=self.radio_sigma, value=1, indicator=0, width=12)
        self.button_sigma.grid(column=8, row=0, columnspan=1, sticky='e', padx=2, pady=0)
        self.entry_sigma = tk.Entry(self.frame_control, width=6, justify='center')
        self.entry_sigma.insert(0, 3)
        self.entry_sigma.grid(row=0, column=9, sticky='w', padx=4, pady=0)
        
        self.button_outlier = tk.Radiobutton(self.frame_control, text='Mask Outliers', 
                            variable=self.radio_sigma, value=2, indicator=0, width=12)  
        self.button_outlier.grid(column=8, row=1, columnspan=1, sticky='e', padx=2, pady=0)
        self.entry_outlier = tk.Entry(self.frame_control, width=6, justify='center')
        self.entry_outlier.insert(0, 0)
        self.entry_outlier.grid(row=1, column=9, sticky='w', padx=4, pady=0)

        self.variable1 = tk.StringVar()
        self.so1 = ttk.Combobox(self.frame_control, textvariable=self.variable1, justify='center')
        self.so1.config(values='0', width=5)
        self.so1.grid(column=4, row=0,  sticky='w', padx=4, pady=0)
        self.so1.option_add('*TCombobox*Listbox.Justify', 'center')
        self.label_combo1 = tk.Label(self.frame_control, text="A")
        self.label_combo1.grid(row=0, column=3, sticky='e', padx=0, pady=0)

        self.variable2 = tk.StringVar()
        self.so2 = ttk.Combobox(self.frame_control, textvariable=self.variable2, justify='center')
        self.so2.config(values='0', width=5)
        self.so2.grid(column=6, row=0,  sticky='w', padx=4, pady=0)
        self.so2.option_add('*TCombobox*Listbox.Justify', 'center')
        self.label_combo2 = tk.Label(self.frame_control, text="B")
        self.label_combo2.grid(row=0, column=5, sticky='e', padx=0, pady=0)

        option_calist = ["A-B", "B-A", "(A-B)/t", "(B-A)/t", "A"]
        self.variable_cal = tk.StringVar()
        self.variable_cal.set("A-B")
        self.option_calbtn = tk.OptionMenu(self.frame_control, self.variable_cal, *option_calist)
        self.option_calbtn.config(width=5)
        self.option_calbtn.grid(column=6, row=1, columnspan=2, sticky='w', padx=0, pady=0)

        self.entry_run_time = tk.Entry(self.frame_control, width=6, justify='center')
        self.entry_run_time.insert(0, 60)
        self.entry_run_time.grid(row=2, column=6, sticky='w', padx=2, pady=0)
        self.label_run_time = tk.Label(self.frame_control, text="Time")
        self.label_run_time.grid(row=2, column=5, sticky='e', padx=2, pady=0)

        self.label_unit = tk.Label(self.frame_control, text="Unit", anchor="e")
        self.label_unit.grid(row=2, column=11, sticky='e', padx=2, pady=0)
        self.entry_unit = tk.Entry(self.frame_control, width=6, justify='center')
        self.entry_unit.insert(0, 'Å')
        self.entry_unit.grid(row=2, column=12, sticky='w', padx=2, pady=0)      

        self.label_contour = tk.Label(self.frame_control, width=8, text="Contours", anchor="e")
        self.label_contour.grid(row=0, column=11, sticky='e', padx=2, pady=0)
        self.entry_contour = tk.Entry(self.frame_control, width=6, justify='center')
        self.entry_contour.insert(0, 10)
        self.entry_contour.grid(row=0, column=12, sticky='w', padx=2, pady=0)

        option_list3 = ["Value", "Sign", "Dot", "None"]
        self.variable3 = tk.StringVar()
        self.variable3.set("Value")
        self.option_button3 = tk.OptionMenu(self.frame_control, self.variable3, *option_list3)
        self.option_button3.config(width=5)
        self.option_button3.grid(column=12, row=1,  sticky='w', padx=1, pady=0)
        self.marker = tk.Label(self.frame_control, width=8, text="Markers", anchor="e")
        self.marker.grid(row=1, column=11, sticky='e', padx=2, pady=0)

        self.spacer1 = tk.Label(self.frame_control, width=9, text="")        
        self.spacer1.grid(row=0, column=2, sticky='ew')
        self.spacer2 = tk.Label(self.frame_control, width=12, text="")        
        self.spacer2.grid(row=0, column=7, sticky='ew')
        self.spacer3 = tk.Label(self.frame_control, width=8, text="")        
        self.spacer3.grid(row=1, column=10, sticky='ew')
        self.spacer4 = tk.Label(self.frame_control, width=10, text="")        
        self.spacer4.grid(row=1, column=13, sticky='ew')
        self.spacer5 = tk.Label(self.frame_control, width=0, text="")        
        self.spacer5.grid(row=1, column=17, sticky='ew')
        self.spacer6 = tk.Label(self.frame_control, width=6, text="")        
        self.spacer6.grid(row=1, column=20, sticky='ew')

        self.rotation_entry = tk.Entry(self.frame_control, width=6, justify='center')
        self.rotation_entry.insert(0, 0)
        self.rotation_entry.grid(row=2, column=9, sticky='w', padx=4, pady=0)
        self.label_rotation = tk.Label(self.frame_control, width=10, text="Set Rotation", anchor="e")
        self.label_rotation.grid(row=2, column=8, sticky='e', padx=0, pady=0)

        self.var_animated = tk.IntVar()
        self.check_animated = tk.Checkbutton(self.frame_control, text='3D Animated',
                            variable=self.var_animated)
        self.check_animated.grid(row=2, column=15, sticky='e', padx=10, pady=0, columnspan=2)
        self.var_decom = tk.IntVar()
        self.check_decom = tk.Checkbutton(self.frame_control, text='Decompose',
                            variable=self.var_decom)
        self.check_decom.grid(row=2, column=14, sticky='w', padx=0, pady=0, columnspan=2)

        self.var_limits = tk.IntVar()
        self.var_limits.set(2)
        self.button_range = tk.Radiobutton(self.frame_control, text='Auto Range', variable=self.var_limits,
                            value=2, indicator=0, width=11)  
        self.button_range.grid(column=14, row=0, columnspan=1, sticky='e', padx=4, pady=0)
        self.button_ul = tk.Radiobutton(self.frame_control, text='Set Range', variable=self.var_limits, 
                            value=1, indicator=0, width=11)
        self.button_ul.grid(column=14, row=1, columnspan=1, sticky='e', padx=4, pady=0)
        self.limits_entry = tk.Entry(self.frame_control, width=7, justify='center')
        self.limits_entry.insert(0, '1000')
        self.limits_entry.grid(row=1, column=15, sticky='w', padx=2, pady=0)
        
        self.button_plot = tk.Button(self.frame_control, text="Plot", command=self.plot_graph, width=12)
        self.button_plot.grid(row=0, column=21, sticky="e", padx=4, pady=6)
        self.button_ps = tk.Button(self.frame_control, text="Plot + Save", command=self.save_file,
                            width=12)
        self.button_ps.grid(row=1, column=21, sticky="e", padx=4, pady=0)
        self.button_copy = tk.Button(self.frame_control, text="Copy Map", command=self.copy_clipboard,
                            width=12)
        self.button_copy.grid(row=2, column=21, sticky="e", padx=4, pady=6)
        
        self.canvas_graph = tk.Canvas(self.frame_display, bg='white', bd=0, scrollregion=[0,0,1650,1400])
        self.canvas_graph.grid(row=0, column=0, sticky="nesw")
        self.canvas_graph.columnconfigure(0, weight=1)
        self.canvas_graph.rowconfigure(0, weight=1)

        self.scrollbar_vertical = tk.Scrollbar(self.frame_display, orient="vertical", relief=tk.SUNKEN, bd=0,
                            width=16)
        self.scrollbar_vertical.grid(row=0, column=1, sticky="ns")
        self.canvas_graph.configure(yscrollcommand=self.scrollbar_vertical.set)
        self.scrollbar_vertical.config(command=self.canvas_graph.yview)

        self.scrollbar_horizontal = tk.Scrollbar(self.frame_display, orient="horizontal", relief=tk.SUNKEN,
                            bd=0, width=16)
        self.scrollbar_horizontal.grid(row=1, column=0, sticky="ew")
        self.canvas_graph.configure(xscrollcommand=self.scrollbar_horizontal.set)
        self.scrollbar_horizontal.config(command=self.canvas_graph.xview)
        
        tk.mainloop()

    def open_atlas(self):
        global df1
        global df3
        global df32
        global col2_list

        df5 = pd.DataFrame()

        filepath = askopenfilename(filetypes=[("CSV", "*.csv"), ("Excel", "*.xlsx"), ("Excel", "*.xls"),
                            ("All", "*.*")])
        path = filepath
        self.file_path.configure(text=path)
        filename = os.path.basename(filepath)
        
        extension = os.path.splitext(filename)[1]
        if (extension == '.csv'):
            def csv_to_excel(csv_file, excel_file):
                csv_data = []
                with open(csv_file) as file_obj:
                    reader = csv.reader(file_obj)
                    for row in reader:
                        csv_data.append(row)
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                for row in csv_data:
                    sheet.append(row)
                workbook.save(excel_file)
            if __name__ == "__main__":
                csv_to_excel(filepath, './tmp.xlsx')
                df_data = pd.read_excel('./tmp.xlsx')
                os.remove('tmp.xlsx')
        else:
            df_data = pd.read_excel(filepath)
            
        df1 = pd.DataFrame(df_data)
        cols = len(df1.axes[1]) +1
        column_names = pd.Series(np.arange(1, cols, 1))
        
        mask = np.column_stack([df1[col].str.contains(r"Wafer ID:", na=False) for col in df1])
        df2 = df1.loc[mask.any(axis=1)]
        df2.columns = column_names
        df3 = (df2.loc[:, 2]).reset_index()
        col2_list = df3[2].values.tolist()
        
        mask2 = np.column_stack([df1[col].str.contains(r"Max", na=False) for col in df1])
        df22 = df1.loc[mask2.any(axis=1)]
        df22.columns = column_names
        df32 = (df22.loc[:, 2]).reset_index()

        variable1 = tk.StringVar()
        self.cb1 = ttk.Combobox(self.frame_control, textvariable=variable1, justify='center', height=40)
        self.cb1.bind('<<ComboboxSelected>>', self.callback1)
        self.cb1.grid(column=4, row=0,  sticky='w', padx=4, pady=0)
        self.cb1.config(values=col2_list, width=5)
        self.cb1['state']= 'readonly'

        variable2 = tk.StringVar()
        self.cb2 = ttk.Combobox(self.frame_control, textvariable=variable2, justify='center', height=40)
        self.cb2.bind('<<ComboboxSelected>>', self.callback2)
        self.cb2.grid(column=6, row=0,  sticky='w', padx=4, pady=0)
        self.cb2.config(values=col2_list, width=5)
        self.cb2['state']= 'readonly'
        
    def callback1(self, *args):
        global dfA1
        global dfX1
        global dfY1
        global start1
        global end1
        global col_num1
        global film1

        position_a = str(self.cb1.current())
        df4 = df3.iloc[[position_a]]
        row_num1 = df4['index'].loc[df4.index[0]]
        start1 = row_num1 + 5
        df421 = df32.iloc[[position_a]]
        row_num12 = df421['index'].loc[df421.index[0]]
        end1 = row_num12 - 1

        if (df1.iloc[(start1-1), 2] == 'Resist (Å)'):
            film1 = 'Resist'
            col_num1 = 2
        elif (df1.iloc[(start1-1), 3] == 'Si3N4 (Å)'):
            film1 = 'Si3N4'
            col_num1 = 3
        elif (df1.iloc[(start1-1), 3] == 'Poly (Å)'):
            film1 = 'Poly'
            col_num1 = 3
        elif (df1.iloc[(start1-1), 2] == 'SiO2 (Å)'):
            film1 = 'SiO2'
            col_num1 = 2
        elif (df1.iloc[(start1-1), 2] == 'AC (Å)'):
            film1 = 'AC'
            col_num1 = 2
        else:
            pass
        
        dfA = df1.iloc[start1:end1, col_num1]
        dfA1 = dfA.reset_index()
        dfA1.columns = ['rows_a', 'A']
        row_x1 = row_num1 + 4
        dfx = df1.iloc[row_x1]
        col_ct_x = dfx.count() - 2
        col_ct_y = dfx.count() - 1

        dfX = df1.iloc[start1:end1, col_ct_x]
        dfX1 = dfX.reset_index()
        dfX1.columns = ['rows_x', 'X']
        dfY = df1.iloc[start1:end1, col_ct_y]
        dfY1 = dfY.reset_index()
        dfY1.columns = ['rows_y', 'Y']

    def callback2(self, *args):
        global df5
        global start2
        global end2
        
        position_b = str(self.cb2.current())
        df4 = df3.iloc[[position_b]]
        row_num2 = df4['index'].loc[df4.index[0]]
        start2 = row_num2 + 5
        df422 = df32.iloc[[position_b]]
        row_num22 = df422['index'].loc[df422.index[0]]
        end2 = row_num22 - 1
        
        dfB = df1.iloc[start2:end2, col_num1]
        dfB1 = dfB.reset_index()      
        dfB1.columns = ['rows_b', 'B']
        df5 = pd.concat([dfA1, dfB1, dfX1, dfY1], axis=1)

        if df5['A'].equals(df5['B']):
            import win32api
            win32api.MessageBox(0, 'You have selected the same data set for both "A" and "B"', 'Error:')
        else:
            pass
        
        tk.mainloop()

    def open_workbook(self):
        global df1

        df5 = pd.DataFrame()

        filepath = askopenfilename(filetypes=[("CSV", "*.csv"), ("Excel", "*.xlsx"), ("Excel", "*.xls"),
                            ("All", "*.*")])
        path = filepath
        self.file_path.configure(text=path)
        filename = os.path.basename(filepath)
        
        extension = os.path.splitext(filename)[1]
        if (extension == '.csv'):
            def csv_to_excel(csv_file, excel_file):
                csv_data = []
                with open(csv_file) as file_obj:
                    reader = csv.reader(file_obj)
                    for row in reader:
                        csv_data.append(row)
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                for row in csv_data:
                    sheet.append(row)
                workbook.save(excel_file)
            if __name__ == "__main__":
                csv_to_excel(filepath, './tmp.xlsx')
                df_data = pd.read_excel('./tmp.xlsx')
                os.remove('tmp.xlsx')
        else:
            df_data = pd.read_excel(filepath)

        df1 = pd.DataFrame(df_data)
        nan_value = float("NaN")
        df1.replace("", nan_value, inplace=True)
        df1.dropna(how='all', axis=1, inplace=True)        
        col_name = list(df1.columns.values)
        col_name1 = [s.replace('Unnamed:', '') for s in col_name]

        def make_int(s):
            if not s:
                return s
            try:
                f = float(s)
                i = int(f)
                return i if f == i else f
            except ValueError:
                return s
        converted = list(map(make_int, col_name1))
        converted = [t + 1 if type(t) in [int, float, bool] else t for t in converted]
        del converted[0:2]

        output_list = []
        for element in converted:
            value = str(element)
            output_list.append(value)

        prefix='Col '
        pre_res = []
        for item in output_list:
            pre_res.append(prefix + item)

        variable3 = tk.StringVar()
        self.cb3 = ttk.Combobox(self.frame_control, textvariable=variable3, justify='center', height=40)
        self.cb3.bind('<<ComboboxSelected>>', self.callback3)
        self.cb3.grid(column=4, row=0,  sticky='w', padx=4, pady=0)
        self.cb3.config(values=pre_res, width=5)
        self.cb3['state']= 'readonly'

        variable4 = tk.StringVar()
        self.cb4 = ttk.Combobox(self.frame_control, textvariable=variable4, justify='center', height=40)
        self.cb4.bind('<<ComboboxSelected>>', self.callback4)
        self.cb4.grid(column=6, row=0,  sticky='w', padx=4, pady=0)
        self.cb4.config(values=pre_res, width=5)
        self.cb4['state']= 'readonly'

    def callback3(self, *args):
        global end3
        global film1
        global dfA

        film1 = ''
        position_a = int(str(self.cb3.current())) +2
        start3 = 1
        row_num3 = len(df1.index)
        end3 = row_num3 +1
        dfA = df1.iloc[0:end3, position_a]

    def callback4(self, *args):
        global df5

        position_b = int(str(self.cb4.current())) +2
        dfB = df1.iloc[0:end3, position_b]
        dfX = df1.iloc[0:end3, 0]
        dfY = df1.iloc[0:end3, 1]
        df5 = pd.concat([dfX, dfY, dfA, dfB], axis=1)
        df5.columns = ['X', 'Y', 'A', 'B']

        if df5['A'].equals(df5['B']):
            import win32api
            win32api.MessageBox(0, 'You have selected the same data set for both "A" and "B"', 'Error:')
        else:
            pass

        tk.mainloop()

    def plot_graph(self):
        global id
        global df
        global unit
        global N1
        
        self.canvas_graph.delete('all')       
        id = self.entry_id.get()
        ID = id
        unit = self.entry_unit.get()

        df5['X'] = df5['X'].apply(float)
        df5['Y'] = df5['Y'].apply(float)
            
        value1 = self.variable_cal.get()
        if (value1 == 'A-B'):
            df5['A'] = df5['A'].apply(float)
            df5['B'] = df5['B'].apply(float)
            df5["Z"] = df5["A"] - df5["B"]
            z = df5["Z"]
            data = [df5["X"], df5["Y"], z]
            headers = ["A", "B", "Z"]
            orignal_df = pd.concat(data, axis=1, keys=headers)
            
        elif (value1 == '(A-B)/t'):
            df5['A'] = df5['A'].apply(float)
            df5['B'] = df5['B'].apply(float)
            df5["Z"] = df5["A"] - df5["B"]
            z = df5["Z"]
            data = [df5["X"], df5["Y"], z]
            headers = ["A", "B", "Z"]
            orignal_df = pd.concat(data, axis=1, keys=headers)
            
        elif (value1 == 'B-A'):
            df5['A'] = df5['A'].apply(float)
            df5['B'] = df5['B'].apply(float)
            df5["Z"] = - df5["A"] + df5["B"]
            z = df5["Z"]
            data = [df5["X"], df5["Y"], z]
            headers = ["A", "B", "Z"]
            orignal_df = pd.concat(data, axis=1, keys=headers)
            
        elif (value1 == '(B-A)/t'):
            df5['A'] = df5['A'].apply(float)
            df5['B'] = df5['B'].apply(float)
            df5["Z"] = - df5["A"] + df5["B"]
            z = df5["Z"]
            data = [df5["X"], df5["Y"], z]
            headers = ["A", "B", "Z"]
            orignal_df = pd.concat(data, axis=1, keys=headers)
            
        elif (value1 == 'A'):
            df5['A'] = df5['A'].apply(float)
            df5['C'] = 0
            df5['C'] = df5['C'].apply(float)
            df5["Z"] = df5["A"] - df5["C"]
            z = df5["Z"]
            data = [df5["X"], df5["Y"], z]
            headers = ["A", "B", "Z"]
            orignal_df = pd.concat(data, axis=1, keys=headers)
        else:
            pass

        value = self.radio_sigma.get()
        if (value == 1):
            sigma_var = float(self.entry_sigma.get())
            z_scores = np.abs(stats.zscore(orignal_df))
            df = orignal_df[(z_scores < sigma_var).all(axis=1)]
            N1 = str(orignal_df.shape[0] - df.shape[0])
        else:
            N2 = int(self.entry_outlier.get())
            orignal_df["dist"] = np.abs(orignal_df["Z"]  - np.mean(orignal_df["Z"]))
            sort_df = orignal_df.sort_values(by="dist", ascending=False)
            df = sort_df.iloc[N2:]
            Z = df["Z"]
            stdev = np.std(Z)
            top = max(df["dist"])
            sigma2 = round(top/stdev, 2)

# 2d contour
        fig1 = plt.figure(figsize=(2.5, 2.5))
        spec1 = gridspec.GridSpec(ncols=1, nrows=1)
        ax1 = fig1.add_subplot(spec1[0, 0])
        
        degree_theta = float(self.rotation_entry.get())
        radian_theta = math.radians(degree_theta)
        if (degree_theta == 0):
            x2d = df["A"]
            y2d = df["B"]
        else:
            x_original = df["A"]
            y_original = df["B"]        
            x2d = math.cos(radian_theta)*x_original + math.sin(radian_theta)*y_original
            y2d = - math.sin(radian_theta)*x_original + math.cos(radian_theta)*y_original
        z2d = df["Z"]
        
        x_grid = np.linspace(np.min(x2d), np.max(x2d), 300)
        y_grid = np.linspace(np.min(y2d), np.max(y2d), 300)
        X2d, Y2d = np.meshgrid(x_grid, y_grid)
        Z2d = griddata((x2d, y2d), z2d, (X2d, Y2d), method="cubic")
        
        Ave = round(np.mean(z2d), 1)
        Std = np.std(z2d)
        Std_percent = round(100*Std/abs(Ave), 2)
        Max = round(max(z2d), 1)
        Min = round(min(z2d), 1)
        Nonu = round(0.5*100*(Max-Min)/abs(Ave), 2)
        Range = round(max(z2d) - min(z2d), 1)

        value1 = self.variable_cal.get()
        if (value1 == '(A-B)/t'):
            time_pre = self.entry_run_time.get()
            time = float(time_pre)
            Rate_pre = abs(Ave)*(60/time)
            Rate = round(Rate_pre, 1)         
            plt.text(-75,-330, 'Rate (' + unit + '/min)')
            plt.text(45,-330, Rate)
        elif (value1 == '(B-A)/t'):
            time_pre = self.entry_run_time.get()
            time = float(time_pre)
            Rate_pre = abs(Ave)*(60/time)
            Rate = round(Rate_pre, 1)         
            plt.text(-75,-330, 'Rate (' + unit + '/min)')
            plt.text(45,-330, Rate)
        else:
            pass
        
        plt.text(-75,-240, 'Mean (' + unit + ')')
        plt.text(-75,-255, 'NonU (%)')
        plt.text(-75,-270, 'StdD (%)')
        plt.text(-75,-285, 'Max (' + unit + ')')
        plt.text(-75,-300, 'Min (' + unit + ')')
        plt.text(-75,-315, 'Range (' + unit + ')')
        plt.text(45,-240, Ave)
        plt.text(45,-255, Nonu)                                      
        plt.text(45,-270, Std_percent)
        plt.text(45,-285, Max)
        plt.text(45,-300, Min)
        plt.text(45,-315, Range)

        check_box = self.var_limits.get() 
        if (check_box == 1):
            limits_bt_pre = self.limits_entry.get()
            limits_bt = abs(float(limits_bt_pre))
            limits_b = Ave - limits_bt/2
            limits_t = Ave + limits_bt/2
            num = int(self.entry_contour.get())
            if (num >100):
                num = 100
                levels = np.linspace(limits_b, limits_t, num)
            else:
                levels = np.linspace(limits_b, limits_t, num)
            cp = plt.contourf(X2d, Y2d, Z2d, levels = levels, cmap=plt.cm.turbo, alpha=0.95)
            fig1.colorbar(cp, ax=ax1, shrink=0.5, orientation="vertical")
        else:
            contour = int(self.entry_contour.get()) 
            if (contour >100):
                contour = 100
            else:
                pass
            cp = plt.contourf(X2d, Y2d, Z2d, contour, cmap=plt.cm.turbo, alpha=0.95)
            cbar1 = fig1.colorbar(cp, ax=ax1, shrink=0.5, orientation="vertical")
            cbar1.ax.locator_params(nbins=6)

        value = self.radio_sigma.get()
        if (value == 1):
            plt.text(-75,-345, 'Points removed')
            plt.text(45,-345, N1)
        else:
            plt.text(-75,-345, 'Sigma:')
            plt.text(45,-345, sigma2)
        
        circ = Circle((0, 0), 150, facecolor='None', edgecolor='black', lw=0.2, alpha=0.4)
        ax1.add_patch(circ)
        ax1.set_aspect('equal', adjustable='box')

        var3 = self.variable3.get()
        if (var3 == "Value"):
            for x2d, y2d, z2d in zip(x2d, y2d, z2d):
                label = "{:.0f}".format(z2d)
                plt.annotate(label, (x2d, y2d), textcoords="offset points", xytext=(0, -3), ha='center',
                             fontsize=4.3, alpha=0.85)
                ax1.scatter(x2d, y2d, marker='o', s=1, color='k', alpha=0.0001)
        elif (var3 == "Dot"):
            for x2d, y2d, z2d in zip(x2d, y2d, z2d):
                ax1.scatter(x2d, y2d, marker='o', s=1, color='k', alpha=0.3)
        elif (var3 == "None"):
            for x2d, y2d, z2d in zip(x2d, y2d, z2d):
                label = "{:.0f}".format(z2d)
                ax1.scatter(x2d, y2d, marker='o', s=1, color='k', alpha=0.0001)
        else:
            for x2d, y2d, z2d in zip(x2d, y2d, z2d):
                label = "{:.0f}".format(z2d)
                if (z2d > Ave):
                    ax1.scatter(x2d, y2d, marker='+', s=13, linewidths=0.3, color='k', alpha=0.7)
                else:
                    ax1.scatter(x2d, y2d, marker='_', s=13, linewidths=0.3, color='k', alpha=0.7)

        ax1.set_xlabel('X (mm)')
        ax1.set_ylabel('Y (mm)')    
        ax1.set_title(id + '    ' + film1)
    
        fig1.tight_layout(pad=0.1)
        plot_id1 = "contour"
        plt.savefig(plot_id1, bbox_inches='tight')        
        img1 = Image.open("contour.png")
        plot_id1_resized = img1.resize((350, 380))
        plot_id1_resized.save("contour_resized.png")     
        img11 = ImageTk.PhotoImage(Image.open("contour_resized.png"))
        self.canvas_graph.create_image(20, 35, anchor="nw", image=img11)
        os.remove("contour_resized.png")
        plt.clf()
        plt.close(fig1)
            
# cross_section 1
        fig_crs1 = plt.figure(figsize=(2.3, 1.8))     
        spec_crs1 = gridspec.GridSpec(ncols=1, nrows=1)
        ax_crs1 = fig_crs1.add_subplot(spec_crs1[0, 0])
        ax_crs1.set_xlabel('Cross section (mm)')
        ax_crs1.set_ylabel(unit)
        ax_crs1.set(title=id)

        degree_theta = float(self.rotation_entry.get())
        radian_theta = math.radians(degree_theta)
        if (degree_theta == 0):
            x_crs1 = df["A"]
            y_crs1 = df["B"]
        else:
            x_oric1 = df["A"]
            y_oric1 = df["B"]
            x_crs1 = math.cos(radian_theta)*x_oric1+ math.sin(radian_theta)*y_oric1
            y_crs1 = - math.sin(radian_theta)*x_oric1 + math.cos(radian_theta)*y_oric1
        z_crs1 = df["Z"]

        n_crs1 = 80
        x_grid_crs1 = np.linspace(np.min(x_crs1), np.max(x_crs1), n_crs1)
        y_grid_crs1 = np.linspace(np.min(y_crs1), np.max(y_crs1), n_crs1)
        X_crs1, Y_crs1 = np.meshgrid(x_grid_crs1, y_grid_crs1)
        
        Y_crs1 = 0
        Z_crs1 = griddata((x_crs1, y_crs1), z_crs1, (X_crs1, Y_crs1), method="cubic")

        degree_crs1_90 = 90
        radian_crs1_90 = math.radians(degree_crs1_90)     
        x_crs1_90 = math.cos(radian_crs1_90)*x_crs1 + math.sin(radian_crs1_90)*y_crs1
        y_crs1_90 = - math.sin(radian_crs1_90)*x_crs1 + math.cos(radian_crs1_90)*y_crs1

        Y_crs1 = 0
        Z_crs1_90 = griddata((x_crs1_90, y_crs1_90), z_crs1, (X_crs1, Y_crs1), method="cubic")

        ave1 = np.mean(z_crs1)
        count_row = n_crs1
        Std = np.std(z_crs1)
        sigma_var = float(self.entry_sigma.get())

        Z_crs1_1 = [ave1 - 3*Std]*count_row 
        Z_crs1_2 = [ave1]*count_row 
        Z_crs1_3 = [ave1 + 3*Std]*count_row

        ax_crs1.plot(X_crs1, Z_crs1_1, '.', color='grey', markersize=0.015, alpha=0.165)
        ax_crs1.plot(X_crs1, Z_crs1_2, '.', color='grey', markersize=0.015, alpha=0.165)
        ax_crs1.plot(X_crs1, Z_crs1_3, '.', color='grey', markersize=0.015, alpha=0.165)

        ax_crs1.plot(X_crs1, Z_crs1, 'o', color='red', markersize=0.01)
        ax_crs1.plot(X_crs1, Z_crs1_90, 'o', color='blue', markersize=0.01)
        
        ax_crs1.legend(("Red: X-X", "Blue: Y-Y"), frameon=False, loc='upper right')
        top_x_crs1 = max(x_crs1)
        btn_x_crs1 = min(x_crs1)
        ax_crs1.set_xlim(btn_x_crs1, top_x_crs1)

        check_box = self.var_limits.get() 
        if (check_box == 1):
            limits_y_pre = self.limits_entry.get()
            limits_y = abs(float(limits_y_pre))
            btn_y_crs1 = Ave - limits_y/2
            top_y_crs1 = Ave + limits_y/2
            ax_crs1.set_ylim(btn_y_crs1, top_y_crs1)            
        else:
            top_y_crs1 = Ave + Range*1.8
            btn_y_crs1 = Ave - Range*1.8   
            ax_crs1.set_ylim(btn_y_crs1, top_y_crs1)

        fig_crs1.tight_layout(pad=1)    
        plot_id_crs1 ="cross_section1"
        plt.savefig(plot_id_crs1, bbox_inches='tight')
        img_crs1 = Image.open("cross_section1.png")
        plot_crs1_resized = img_crs1.resize((347, 247))
        plot_crs1_resized.save("cross_resized1.png")     
        img_crs1 = ImageTk.PhotoImage(Image.open("cross_resized1.png"))
        self.canvas_graph.create_image(0, 420, anchor="nw", image=img_crs1)
        os.remove("cross_resized1.png")
        plt.clf()
        plt.close(fig_crs1)

# cross_section 2
        fig_crs2 = plt.figure(figsize=(2.3, 1.8))     
        spec_crs2 = gridspec.GridSpec(ncols=1, nrows=1)
        ax_crs2 = fig_crs2.add_subplot(spec_crs2[0, 0])
        ax_crs2.set_xlabel('Cross section (mm)')
        ax_crs2.set_ylabel(unit)
        ax_crs2.set(title=id)

        degree_theta = float(self.rotation_entry.get())
        radian_theta = math.radians(degree_theta)
        if (degree_theta == 0):
            x_crs2 = df["A"]
            y_crs2 = df["B"]
        else:
            x_oric2 = df["A"]
            y_oric2 = df["B"]
            x_crs2 = math.cos(radian_theta)*x_oric2+ math.sin(radian_theta)*y_oric2
            y_crs2 = - math.sin(radian_theta)*x_oric2 + math.cos(radian_theta)*y_oric2
        z_crs2 = df["Z"]

        x_grid_crs2 = np.linspace(np.min(x_crs2), np.max(x_crs2), 80)
        y_grid_crs2 = np.linspace(np.min(y_crs2), np.max(y_crs2), 80)
        X_crs2, Y_crs2 = np.meshgrid(x_grid_crs2, y_grid_crs2)

        Y_max = max(y_crs2)
        Y_min = min(y_crs2)
        ran_crs2 = (max(y_crs2) - min(y_crs2))/4
        Y_cros1 = - ran_crs2
        Y_cros3 = 0
        Y_cros5 = ran_crs2

        Z_crs2_1 = griddata((x_crs2, y_crs2), z_crs2, (X_crs2, Y_cros1), method="cubic")
        Z_crs2_3 = griddata((x_crs2, y_crs2), z_crs2, (X_crs2, Y_cros3), method="cubic")
        Z_crs2_5 = griddata((x_crs2, y_crs2), z_crs2, (X_crs2, Y_cros5), method="cubic")

        ax_crs2.plot(X_crs2, Z_crs1_1, '.', color='grey', markersize=0.015, alpha=0.165)
        ax_crs2.plot(X_crs2, Z_crs1_2, '.', color='grey', markersize=0.015, alpha=0.165)
        ax_crs2.plot(X_crs2, Z_crs1_3, '.', color='grey', markersize=0.015, alpha=0.165)

        ax_crs2.plot(X_crs2, Z_crs2_5, 'o', color='orange', markersize=0.01)      
        ax_crs2.plot(X_crs2, Z_crs2_3, 'o', color='red', markersize=0.01)
        ax_crs2.plot(X_crs2, Z_crs2_1, 'o', color='green', markersize=0.01)
           
        ax_crs2.legend(("Green: 1st X-X", "Red: 2nd X-X", "Yellow: 3rd X-X"), frameon=False,
                                loc='upper right')
        top_x_crs2 = max(x_crs2)
        btn_x_crs2 = min(x_crs2)
        ax_crs2.set_xlim(btn_x_crs2, top_x_crs2)

        check_box = self.var_limits.get() 
        if (check_box == 1):
            limits_y_pre = self.limits_entry.get()
            limits_y = abs(float(limits_y_pre))
            btn_y_crs2 = Ave - limits_y/2
            top_y_crs2 = Ave + limits_y/2
            ax_crs2.set_ylim(btn_y_crs2, top_y_crs2)            
        else:
            top_y_crs2 = Ave + Range*1.8
            btn_y_crs2 = Ave - Range*1.8   
            ax_crs2.set_ylim(btn_y_crs2, top_y_crs2)

        fig_crs2.tight_layout(pad=1)    
        plot_id_crs2 ="cross_section2"
        plt.savefig(plot_id_crs2, bbox_inches='tight')
        img_crs2 = Image.open("cross_section2.png")
        plot_crs2_resized = img_crs2.resize((347, 247))
        plot_crs2_resized.save("cross_resized2.png")     
        img_crs2 = ImageTk.PhotoImage(Image.open("cross_resized2.png"))
        self.canvas_graph.create_image(370, 420, anchor="nw", image=img_crs2)
        os.remove("cross_resized2.png")
        plt.clf()
        plt.close(fig_crs2)

# 3d_contour
        fig3d = plt.figure(figsize=(2.5, 2.5))
        ax3d = plt.axes(projection="3d")
        ax3d.view_init(30, 240)
        degree_theta = float(self.rotation_entry.get())
        radian_theta = math.radians(degree_theta)
        if (degree_theta == 0):
            x3d = df["A"]
            y3d = df["B"]
        else:
            degree_theta = float(self.rotation_entry.get())
            radian_theta = math.radians(degree_theta)
            x_orig = df["A"]
            y_orig = df["B"]
            x3d = math.cos(radian_theta)*x_orig + math.sin(radian_theta)*y_orig
            y3d = - math.sin(radian_theta)*x_orig + math.cos(radian_theta)*y_orig
        z3d = df["Z"] 

        x3d_grid = np.linspace(np.min(x3d), np.max(x3d), 120)
        y3d_grid = np.linspace(np.min(y3d), np.max(y3d), 120)
        X3d, Y3d = np.meshgrid(x3d_grid, y3d_grid)
        
        Z3d = griddata((x3d, y3d), z3d, (X3d, Y3d), method="cubic")
        ax3d.contour3D(X3d, Y3d, Z3d, 750, cmap='turbo', alpha=1, antialiased=False)
 
        check_box = self.var_limits.get() 
        if (check_box == 1):
            limits_z_pre = self.limits_entry.get()
            limits_z = abs(float(limits_z_pre))
            btn_limitz3d = np.mean(z3d) - limits_z/2
            top_limitz3d = np.mean(z3d) + limits_z/2
            ax3d.set_zlim(btn_limitz3d, top_limitz3d)          
        else:
            top_limitz3d = np.mean(z3d) + (max(z3d)-min(z3d))*1.5
            btn_limitz3d = np.mean(z3d) - (max(z3d)-min(z3d))*4.5
            ax3d.set_zlim(btn_limitz3d, top_limitz3d)

        cset = ax3d.contour(X3d, Y3d, Z3d, 10, zdir='z', offset=btn_limitz3d, linewidths=1,
                        cmap=cm.turbo, alpha=0.9)
            
        ax3d.set_xlabel('X (mm)')
        ax3d.set_ylabel('Y (mm)')
        ax3d.set_zlabel(unit)
        ax3d.set_title('Original 3d map \n Mean = %.1f Å, NonU = %.2f pct' %(Ave, Nonu), fontsize=6)
        
        fig3d.tight_layout(pad=1)    
        plot_id3d ="3d"
        plt.savefig(plot_id3d, bbox_inches='tight')
        img3d = Image.open("3d.png")
        plot_id3d_resized = img3d.resize((356, 356))
        plot_id3d_resized.save("3d_resized.png")     
        img33d = ImageTk.PhotoImage(Image.open("3d_resized.png"))
        self.canvas_graph.create_image(375, 0, anchor="nw", image=img33d)
        os.remove("3d_resized.png")
        plt.clf()
        plt.close(fig3d)

# 3d animated
        from matplotlib.animation import FuncAnimation

        check_box3 = self.var_animated.get() 
        if (check_box3 == 1):
            fig_3d_rot = plt.figure()
            fig_3d_rot.set_size_inches(2, 2)
            ax_3d_rot = plt.axes(projection='3d')
            ax_3d_rot.set_facecolor((1, 1, 1))

            check_box = self.var_limits.get() 
            if (check_box == 1):
                limits_z_pre = self.limits_entry.get()
                top_limitz3d_animated = float(limits_z_pre)
                limits_z_pre2 = self.limits_entry.get()
                btn_limitz3d_animated = float(limits_z_pre2)
                ax_3d_rot.set_zlim(btn_limitz3d, top_limitz3d)          
            else:
                top_limitz3d_animated = np.mean(z3d) + (max(z3d)-min(z3d))*1.85
                btn_limitz3d_animated = np.mean(z3d) - (max(z3d)-min(z3d))*1.85
                ax_3d_rot.set_zlim(btn_limitz3d_animated, top_limitz3d_animated)

            ax_3d_rot.set_xlabel('X', fontsize=4)
            ax_3d_rot.set_ylabel('Y', fontsize=4)
            ax_3d_rot.set_zlabel('Z', fontsize=4)
            ax_3d_rot.set_title(id, fontsize=5)

            ax_3d_rot.plot_surface(X3d, Y3d, Z3d, rstride=6, cstride=6, cmap='viridis', alpha=0.85,
                               edgecolor='none')
            plt.axis('on')

            def update(i, fig_3d_rot, ax_3d_rot):
                ax_3d_rot.view_init(elev=30., azim=i)
                return fig_3d_rot, ax_3d_rot

            anim = FuncAnimation(fig_3d_rot, update, frames=np.arange(0, 360, 1), repeat=True,
                                fargs=(fig_3d_rot, ax_3d_rot), interval=100)
            anim.save('rgb_cube.gif', dpi=60, writer='pillow')
            plt.pause(0.1)
            os.remove('rgb_cube.gif')
            #for website
            #anim.save('rgb_cube.gif', dpi=600, writer='pillow') 
            #os.remove('rgb_cube.gif')
        else:
            pass

# plot all
        check_box2 = self.var_decom.get() 
        if (check_box2 == 1):
            
# scatter1
            fig5 = plt.figure(figsize=(2.3, 1.7))
            spec5 = gridspec.GridSpec(ncols=1, nrows=1)

            df["radius (mm)"] = np.sqrt(df["A"]**2 + df["B"]**2)
            df["theta_pre (degree)"] = np.arctan2(df["B"], df["A"])*180/np.pi
            df["theta (degree)"] = df["theta_pre (degree)"].apply(lambda x: x + 360 if x < 0 else x)

            ax5 = fig5.add_subplot(spec5[0, 0])       
            a1 = 1
            b1 = 0.1
            df["u"] = a1*df["radius (mm)"] + b1*df["theta (degree)"]
            df = df.sort_values(by="u")
            df["Point"] = range(1, 1 + len(df["u"]))
            x1 = df["Point"] 
            y1 = df["Z"]

            ave1 = np.mean(y1)
            count_row = df.shape[0]
            Std = np.std(y1)
            sigma_var = float(self.entry_sigma.get())

            y2 = [ave1 - 3*Std]*count_row 
            y3 = [ave1]*count_row 
            y4 = [ave1 + 3*Std]*count_row

            ax5.plot(x1, y1, 'o', markersize=1.5, alpha=0.7)
            ax5.plot(x1, y2, '-', markersize=0.001, c='grey', alpha=0.15)
            ax5.plot(x1, y3, '-', markersize=0.001, c='grey', alpha=0.15)
            ax5.plot(x1, y4, '-', markersize=0.001, c='grey', alpha=0.15)

            ax5.set_xlabel('Measurement point', fontsize=5)
            ax5.set_ylabel(unit, fontsize=5)
            ax5.set(title = 'As radius increasing')

            check_box = self.var_limits.get() 
            if (check_box == 1):
                limit1_bt_pre = self.limits_entry.get()
                limit1_bt = abs(float(limit1_bt_pre))
                btn_limit1 = Ave - limit1_bt/2
                top_limit1 = Ave + limit1_bt/2
                ax5.set_ylim(btn_limit1, top_limit1)
            else:
                top_limit1 = Ave + Range*1.8
                btn_limit1 = Ave - Range*1.8
                ax5.set_ylim(btn_limit1, top_limit1)

            fig5.tight_layout(pad=1)
            plot_id5 ="scatter1"
            plt.savefig(plot_id5, bbox_inches='tight')
            img5 = Image.open("scatter1.png")
            plot_id5_resized = img5.resize((352, 247))
            plot_id5_resized.save("scatter1_resized.png")      
            img5 = ImageTk.PhotoImage(Image.open("scatter1_resized.png")) 
            self.canvas_graph.create_image(0, 690, anchor="nw", image=img5)
            os.remove("scatter1_resized.png")
            plt.clf()
            plt.close(fig5)

# scatter2
            fig6 = plt.figure(figsize=(2.3, 1.7))
            spec6 = gridspec.GridSpec(ncols=1, nrows=1)

            df["radius (mm)"] = np.sqrt(df["A"]**2 + df["B"]**2)
            df["theta_pre (degree)"] = np.arctan2(df["B"], df["A"])*180/np.pi
            df["theta (degree)"] = df["theta_pre (degree)"].apply(lambda x: x + 360 if x < 0 else x)

            ax6 = fig6.add_subplot(spec6[0, 0]) 
            a2 = 0.1
            b2 = 1
            df["v"] = a2*df["radius (mm)"] + b2*df["theta (degree)"]    
            df = df.sort_values(by="v")
            df["Point"] = range(1, 1 + len(df["v"]))
            x2 = df["Point"] 
            y2 = df["Z"] 

            ave1 = np.mean(y1)
            count_row = df.shape[0]
            Std = np.std(y1)
            sigma_var = float(self.entry_sigma.get())

            y2_2 = [ave1 - 3*Std]*count_row 
            y2_3 = [ave1]*count_row 
            y2_4 = [ave1 + 3*Std]*count_row 

            ax6.plot(x2, y2, 'o', markersize=1.5, c='red', alpha=0.7)
            ax6.plot(x2, y2_2, '-', markersize=0.001, c='grey', alpha=0.15)
            ax6.plot(x2, y2_3, '-', markersize=0.001, c='grey', alpha=0.15)
            ax6.plot(x2, y2_4, '-', markersize=0.001, c='grey', alpha=0.15)

            ax6.set_xlabel('Measurement point', fontsize=5)
            ax6.set_ylabel(unit, fontsize=5)
            ax6.set(title='As theta increasing')

            check_box = self.var_limits.get() 
            if (check_box == 1):
                limit2_bt_pre = self.limits_entry.get()
                limit2_bt = abs(float(limit2_bt_pre))
                btn_limit2 = Ave - limit2_bt/2
                top_limit2 = Ave + limit2_bt/2
                ax6.set_ylim(btn_limit2, top_limit2)
            else:
                top_limit2 = Ave + Range*1.8
                btn_limit2 = Ave - Range*1.8
                ax6.set_ylim(btn_limit2, top_limit2)

            fig6.tight_layout(pad=1)
            plot_id6 ="scatter2"
            plt.savefig(plot_id6, bbox_inches='tight')
            img6 = Image.open("scatter2.png")
            plot_id6_resized = img6.resize((352, 247))
            plot_id6_resized.save("scatter2_resized.png")      
            img6 = ImageTk.PhotoImage(Image.open("scatter2_resized.png")) 
            self.canvas_graph.create_image(370, 690, anchor="nw", image=img6)
            os.remove("scatter2_resized.png")
            plt.clf()
            plt.close(fig6)

# histogram
            fig10 = plt.figure(figsize=(2.5, 1.9))
            spec10 = gridspec.GridSpec(ncols=1, nrows=1)       
            ax10 = fig10.add_subplot(spec10[0, 0])
            ax10.set_xlabel(unit, fontsize=6)
            ax10.set_ylabel('Counts', fontsize=6)
            ax10.set(title = 'Histogram')
            z_hist = df["Z"]
            bins = round(len(z_hist)/5)

            z1 = ave1 - 3*Std
            z2 = ave1
            z3 = ave1 + 3*Std

            ax10.hist(z_hist, bins, color='darkblue', alpha=0.5)
            plt.axvline(z1, color='grey', linestyle=':', linewidth=0.5)
            plt.axvline(z2, color='grey', linestyle=':', linewidth=0.5)
            plt.axvline(z3, color='grey', linestyle=':', linewidth=0.5)

            check_box = self.var_limits.get() 
            if (check_box == 1):
                limit_pre = self.limits_entry.get()
                limit_abs = abs(float(limit_pre))
                btn_limit = Ave - limit_abs/2
                top_limit = Ave + limit_abs/2
                ax10.set_xlim((btn_limit, top_limit))
            else:
                btn_limit = Ave - Range*1.3
                top_limit = Ave + Range*1.3
                ax10.set_xlim((btn_limit, top_limit))

            fig10.tight_layout(pad=1)
            plot_id10 ="hist"
            plt.savefig(plot_id10, bbox_inches='tight')
            img10 = Image.open("hist.png")
            plot_id10_resized = img10.resize((316, 242))
            plot_id10_resized.save("hist_resized.png")      
            img10 = ImageTk.PhotoImage(Image.open("hist_resized.png")) 
            self.canvas_graph.create_image(35, 955, anchor="nw", image=img10)
            os.remove("hist_resized.png")
            plt.clf()
            plt.close(fig10)

# cumulative distribution
            fig_cdf = plt.figure(figsize=(2.5, 1.9))
            spec_cdf = gridspec.GridSpec(ncols=1, nrows=1)       
            ax_cdf = fig_cdf.add_subplot(spec_cdf[0, 0])
            ax_cdf.set_xlabel(unit, fontsize=6)
            ax_cdf.set_ylabel('Percent', fontsize=6)
            ax_cdf.set(title = 'Cumulative distribution')

            x_cdf_pre = df["Z"]
            x_cdf = np.sort(x_cdf_pre)
            y_cdf = 100*np.arange(len(x_cdf))/float(len(x_cdf))

            ave = np.mean(x_cdf_pre)
            count_row = df.shape[0]
            Std = np.std(x_cdf_pre)
            sigma_var = float(self.entry_sigma.get())

            x_cdf_2= [ave - 3*Std]*count_row 
            x_cdf_3 = [ave]*count_row 
            x_cdf_4 = [ave + 3*Std]*count_row 

            ax_cdf.plot(x_cdf, y_cdf, 'o', markersize=1.8, color='m', alpha=0.7)
            ax_cdf.plot(x_cdf_2, y_cdf, '-', markersize=0.001, c='grey', alpha=0.15)
            ax_cdf.plot(x_cdf_3, y_cdf, '-', markersize=0.001, c='grey', alpha=0.15)
            ax_cdf.plot(x_cdf_4, y_cdf, '-', markersize=0.001, c='grey', alpha=0.15)

            check_box = self.var_limits.get() 
            if (check_box == 1):
                limit_pre = self.limits_entry.get()
                limit_abs = abs(float(limit_pre))
                btn_limit = Ave - limit_abs/2
                top_limit = Ave + limit_abs/2
                ax_cdf.set_xlim((btn_limit, top_limit))
            else:
                top_cdf = Ave + Range*1.3
                btn_cdf = Ave - Range*1.3
                ax_cdf.set_xlim(btn_cdf, top_cdf)

            fig_cdf.tight_layout(pad=1)
            plot_id_cdf ="cdf"
            plt.savefig(plot_id_cdf, bbox_inches='tight')
            img_cdf = Image.open("cdf.png")
            plot_id_cdf_resized = img_cdf.resize((328, 247))
            plot_id_cdf_resized.save("cdf_resized.png")      
            img_cdf = ImageTk.PhotoImage(Image.open("cdf_resized.png")) 
            self.canvas_graph.create_image(395, 955, anchor="nw", image=img_cdf)
            os.remove("cdf_resized.png")
            plt.clf()
            plt.close(fig_cdf)

# decompose_tilt_corrected
            fig_slp1 = plt.figure(figsize=(2.5, 2.5))
            ax_slp1 = plt.axes(projection="3d")
            degree_theta = float(self.rotation_entry.get())
            radian_rotation = math.radians(degree_theta)
                                    
            if (degree_theta == 0):
                df["x_tc"] = df["A"]
                df["y_tc"] = df["B"]
            else:
                x_orir = df["A"]
                y_orir = df["B"]
                df["x_tc"] = math.cos(radian_rotation)*x_orir + math.sin(radian_rotation)*y_orir
                df["y_tc"] = -math.sin(radian_rotation)*x_orir + math.cos(radian_rotation)*y_orir

            x_slp1 = df["x_tc"]
            y_slp1 = df["y_tc"]
            z_slp1 = df["Z"]

            x_tmp = df[['x_tc', 'y_tc']].values.reshape(-1, 2)
            max_rows = max(len(x_slp1), len(y_slp1))
            x_ones = np.ones([max_rows, 1])
            x = np.append(x_ones, x_tmp, axis=1)
            y_tmp = z_slp1.tolist()
            y = np.array(y_tmp)

            def mse(coef, x, y):
                return np.mean((np.dot(x, coef) - y)**2)/2
            def gradients(coef, x, y):
                return np.mean(x.transpose()*(np.dot(x, coef) - y), axis=1)
            def multilinear_regression(coef, x, y, lr, b1=0.9, b2=0.999, epsilon=1e-8):
                prev_error = 0
                m_coef = np.zeros(coef.shape)
                v_coef = np.zeros(coef.shape)
                moment_m_coef = np.zeros(coef.shape)
                moment_v_coef = np.zeros(coef.shape)
                t = 0
                while True:
                    error = mse(coef, x, y)
                    if abs(error - prev_error) <= epsilon:
                        break
                    prev_error = error
                    grad = gradients(coef, x, y)
                    t += 1
                    m_coef = b1*m_coef + (1-b1)*grad
                    v_coef = b2*v_coef + (1-b2)*grad**2
                    moment_m_coef = m_coef / (1-b1**t)
                    moment_v_coef = v_coef / (1-b2**t)
                    delta = ((lr / moment_v_coef**0.5 + 1e-8)*(b1*moment_m_coef + (1-b1)*grad/(1-b1**t)))
                    coef = np.subtract(coef, delta)
                return coef
 
            coef = np.array([0, 0, 0])
            c = multilinear_regression(coef, x, y, 1e-1)
        
            check_box = self.var_limits.get() 
            if (check_box == 1):
                limits_z_pre = self.limits_entry.get()
                limits_z = abs(float(limits_z_pre))
                btn_limitz_slp1 = np.mean(z_slp1) - limits_z/2
                top_limitz_slp1 = np.mean(z_slp1) + limits_z/2
            else:
                ax_slp1.set_zlim(btn_limitz3d, top_limitz3d)
                btn_limitz_slp1 = np.mean(z_slp1) - (max(z_slp1)-min(z_slp1))*4.5
                top_limitz_slp1 = np.mean(z_slp1) + (max(z_slp1)-min(z_slp1))*1.5

            x_slp1_grid = np.linspace(np.min(x_slp1), np.max(x_slp1), 120)
            y_slp1_grid = np.linspace(np.min(y_slp1), np.max(y_slp1), 120)
            X_slp1, Y_slp1 = np.meshgrid(x_slp1_grid, y_slp1_grid)
            Z_slp1 = griddata((x_slp1, y_slp1), z_slp1, (X_slp1, Y_slp1), method="cubic")
            zz_slp1 = c[1]*x_slp1 + c[2]*y_slp1 + c[0]
            ZZ_slp1 = griddata((x_slp1, y_slp1), zz_slp1, (X_slp1, Y_slp1), method="linear")
            
            ax_slp1.contour3D(X_slp1, Y_slp1, Z_slp1 - ZZ_slp1 + c[0], 450, cmap='turbo', alpha=1,
                                antialiased=False)
            cset = ax_slp1.contour(X_slp1, Y_slp1, Z_slp1 - ZZ_slp1 + c[0], 7, zdir='z', offset=btn_limitz_slp1,
                                linewidths=1, cmap=cm.turbo, alpha=0.9)

            z_after = z_slp1 - zz_slp1 + c[0]
            ave_slp1 = round(np.mean(z_after), 1)
            max_slp1 = round(max(z_after))
            min_slp1 = round(min(z_after))
            nonu_slp1 = round(0.5*100*(max_slp1 - min_slp1)/abs(ave_slp1), 2)

            ax_slp1.set_title('After tilting is corrected \n Mean = %.1f Å, NonU = %.2f pct' %(ave_slp1,
                                nonu_slp1), fontsize=6)
            ax_slp1.set_zlim(btn_limitz_slp1, top_limitz_slp1)
            ax_slp1.view_init(30, 240)
            ax_slp1.set_xlabel('X (mm)')
            ax_slp1.set_ylabel('Y (mm)')
            ax_slp1.set_zlabel(unit)

            plot_slp1 ="slope1"
            plt.savefig(plot_slp1, bbox_inches='tight')
            img_slp1 = Image.open("slope1.png")
            plot_slp1_resized = img_slp1.resize((356, 356))
            plot_slp1_resized.save("slope1_resized.png")     
            img_slp1 = ImageTk.PhotoImage(Image.open("slope1_resized.png"))
            self.canvas_graph.create_image(732, 0, anchor="nw", image=img_slp1)
            os.remove("slope1_resized.png")
            plt.clf()
            plt.close(fig_slp1)
        
# decompose_slope
            fig_slp2 = plt.figure(figsize=(2.5, 2.5))
            ax_slp2 = plt.axes(projection="3d")
            degree_theta = float(self.rotation_entry.get())
            radian_rotation = math.radians(degree_theta)
            
            if (degree_theta == 0):
                x_slp2 = df["A"]
                y_slp2 = df["B"]
            else:
                x_orir = df["A"]
                y_orir = df["B"]
                x_slp2 = math.cos(radian_rotation)*x_orir + math.sin(radian_rotation)*y_orir
                y_slp2 = -math.sin(radian_rotation)*x_orir + math.cos(radian_rotation)*y_orir

            z_slp2 = df["Z"]
            zz_slp2 = c[1]*x_slp2 + c[2]*y_slp2 + c[0]

            actual = y_tmp
            predict = zz_slp2
            correlation_matrix = np.corrcoef(actual, predict)
            corr = correlation_matrix[0, 1]
            R_sq1 = corr**2
            
            v1 = (c[1], c[2], -10**(-7))
            v2 = (0, 0, 10**(-7))
            
            def unit_vector(vector):
                return vector / np.linalg.norm(vector)
            def angle_between(v1, v2):
                v1_u = unit_vector(v1)
                v2_u = unit_vector(v2)
                return np.arccos(np.clip(np.dot(v1_u, v2_u), -1.0, 1.0))
            angle_deg = np.round(180 - math.degrees(angle_between(v1, v2)), decimals=9, out=None)
            angle_deg2 = np.round(90 - angle_deg, decimals=9, out=None) 

            angle_skew = math.degrees(math.atan2(-c[1]/c[2], 1))
            non_acute = angle_skew + 90

            x_slp2_grid = np.linspace(np.min(x_slp2), np.max(x_slp2), 120)
            y_slp2_grid = np.linspace(np.min(y_slp2), np.max(y_slp2), 120)
            X_slp2, Y_slp2 = np.meshgrid(x_slp2_grid, y_slp2_grid)
            Z_slp2 = griddata((x_slp2, y_slp2), zz_slp2, (X_slp2, Y_slp2), method="linear")
                
            btn_limitz_slp2 = btn_limitz3d
            top_limitz_slp2 = top_limitz3d
            Z_slp3 = Z_slp2 - c[0] + np.mean(z3d)
            
            ax_slp2.contour3D(X_slp2, Y_slp2, Z_slp3, 150, cmap='turbo', alpha=0.35, antialiased=False)
            cset = ax_slp2.contour(X_slp2, Y_slp2, Z_slp2, 7, zdir='z', offset=btn_limitz_slp2, linewidths=1,
                                cmap=cm.turbo, alpha=0.9)

            ax_slp2.set_zlim(btn_limitz_slp2, top_limitz_slp2)            
            ax_slp2.view_init(30, 240)
            ax_slp2.set_xlabel('X (mm)')
            ax_slp2.set_ylabel('Y (mm)')
            ax_slp2.set_zlabel(unit)

            ax_slp2.set_title('Abstracted tilting \n z = (%.2f) x + (%.2f) y + (%.2f), R$^2$ = %.2f' % (c[1], c[2],
                                c[0], R_sq1), fontsize=6)
            ax_slp2.text2D(0.2, 0.98, f'Tilt = %.7f°,' % angle_deg2, transform=ax_slp2.transAxes, fontsize=6)
            ax_slp2.text2D(0.53, 0.98, f'Skew = %.2f°' % non_acute, transform=ax_slp2.transAxes,
                                fontsize=6)
            fig_slp2.tight_layout(pad=1)
            
            plot_slp2 ="slope2"
            plt.savefig(plot_slp2, bbox_inches='tight')
            img_slp2 = Image.open("slope2.png")
            plot_slp2_resized = img_slp2.resize((356, 356))
            plot_slp2_resized.save("slope2_resized.png")     
            img_slp2 = ImageTk.PhotoImage(Image.open("slope2_resized.png"))
            self.canvas_graph.create_image(732, 380, anchor="nw", image=img_slp2)
            os.remove("slope2_resized.png")
            plt.clf()
            plt.close(fig_slp2)

# decompose_radius
            from scipy.optimize import curve_fit
            x_rad = df["A"]
            y_rad = df["B"]
            z_rad = df["Z"]
            x = df["radius (mm)"]
            y = z_rad - zz_slp1
            
            def objective(x, a, b, c, d):
                return a*x + b*x**2 + c*x**3 + d 
            popt, _ = curve_fit(objective, x, y)
            a, b, c, d = popt

            fig_rad = plt.figure(figsize=(2.5, 2.5))
            ax_rad = plt.axes(projection="3d")

            zz_rad = a*x + b*x**2 + c*x**3 + d
            Z_rad = griddata((x_rad, y_rad), zz_rad, (X_slp1, Y_slp1), method="cubic")

            my_fitting = np.polyfit(x, y, 3, full=True)
            coeff = my_fitting[0]
            SSE = my_fitting[1][0]
            diff = y - y.mean()
            square_diff = diff ** 2
            SST = square_diff.sum()
            R_sq2 = 1 - SSE/SST 

            btn_limitz_resi= btn_limitz3d
            top_limitz_resi = top_limitz3d

            Z_rad2 = Z_rad + np.mean(z3d) 
            ax_rad.contour3D(X_slp1, Y_slp1, Z_rad2, 450, cmap='turbo', alpha=1, antialiased=False)
            cset = ax_rad.contour(X_slp1, Y_slp1, Z_rad, 7, zdir='z', offset=btn_limitz_resi, linewidths=1,
                                cmap=cm.turbo, alpha=0.9)

            ax_rad.set_zlim(btn_limitz_resi, top_limitz_resi)                                                                                                                      
            ax_rad.view_init(30, 240)      
            ax_rad.set_xlabel('X (mm)')
            ax_rad.set_ylabel('Y (mm)')
            ax_rad.set_zlabel(unit)
            ax_rad.set_title(
                    'Abstracted radial pattern \n z = (%.2f) r + (%.4f) r$^2$ + (%.6f) r$^3$ + (%.2f), R$^2$ = %.2f'
                            % (a, b, c, d, R_sq2), fontsize=6)
                 
            fig_rad.tight_layout(pad=1)    
            plot_rad ="radius"
            plt.savefig(plot_rad, bbox_inches='tight')
            img_rad = Image.open("radius.png")
            plot_rad_resized = img_rad.resize((376, 356))
            plot_rad_resized.save("radius_resized.png")     
            img_rad = ImageTk.PhotoImage(Image.open("radius_resized.png"))  
            self.canvas_graph.create_image(1088, 380, anchor="nw", image=img_rad)
            os.remove("radius_resized.png")
            plt.clf()
            plt.close(fig_rad)

# decompose_residual
            fig_res = plt.figure(figsize=(2.5, 2.5))
            ax_res = plt.axes(projection="3d")
            
            degree_theta = float(self.rotation_entry.get())
            radian_rotation = math.radians(degree_theta)
            if (degree_theta == 0):
                x_res = df["A"]
                y_res = df["B"]
            else:
                x_orire = df["A"]
                y_orire = df["B"]
                x_res = math.cos(radian_rotation)*x_orire + math.sin(radian_rotation)*y_orire
                y_res = -math.sin(radian_rotation)*x_orire + math.cos(radian_rotation)*y_orire

            x_tmp = df[['x_tc', 'y_tc']].values.reshape(-1, 2)
            max_rows = max(len(x_slp1), len(y_slp1))
            x_ones = np.ones([max_rows, 1])
            x = np.append(x_ones, x_tmp, axis=1)
            y_tmp = z_slp1.tolist()
            y = np.array(y_tmp)

            def mse(coef, x, y):
                return np.mean((np.dot(x, coef) - y)**2)/2
            def gradients(coef, x, y):
                return np.mean(x.transpose()*(np.dot(x, coef) - y), axis=1)
            def multilinear_regression(coef, x, y, lr, b1=0.9, b2=0.999, epsilon=1e-8):
                prev_error = 0
                m_coef = np.zeros(coef.shape)
                v_coef = np.zeros(coef.shape)
                moment_m_coef = np.zeros(coef.shape)
                moment_v_coef = np.zeros(coef.shape)
                t = 0
                while True:
                    error = mse(coef, x, y)
                    if abs(error - prev_error) <= epsilon:
                        break
                    prev_error = error
                    grad = gradients(coef, x, y)
                    t += 1
                    m_coef = b1*m_coef + (1-b1)*grad
                    v_coef = b2*v_coef + (1-b2)*grad**2
                    moment_m_coef = m_coef / (1-b1**t)
                    moment_v_coef = v_coef / (1-b2**t)
                    delta = ((lr / moment_v_coef**0.5 + 1e-8)*(b1*moment_m_coef + (1-b1)*grad/(1-b1**t)))
                    coef = np.subtract(coef, delta)
                return coef
            
            coef = np.array([0, 0, 0])
            c = multilinear_regression(coef, x, y, 1e-1)
            
            x_res_grid = np.linspace(np.min(x_res), np.max(x_res), 120)
            y_res_grid = np.linspace(np.min(y_res), np.max(y_res), 120)
            X_res, Y_res = np.meshgrid(x_res_grid, y_res_grid)
            
            z_res = df["Z"]
            zz_res = c[1]*x_res + c[2]*y_res + c[0]
            Z_res = griddata((x_res, y_res), z_res, (X_res, Y_res), method="cubic")
            ZZ_res = griddata((x_res, y_res), zz_res, (X_res, Y_res), method="linear")

            check_box = self.var_limits.get() 
            if (check_box == 1):
                limits_z_pre = self.limits_entry.get()
                limits_z = abs(float(limits_z_pre))
                btn_limitz_res = np.mean(z_res) - limits_z/2
                top_limitz_res = np.mean(z_res) + limits_z/2       
            else:
                top_limitz_res = np.mean(z_res) + (max(z_res)-min(z_res))*1.5
                btn_limitz_res = np.mean(z_res) - (max(z_res)-min(z_res))*4.5
         
            ax_res.contour3D(X_res, Y_res, Z_res - ZZ_res + c[0] - Z_rad, 450, cmap='turbo', alpha=1,
                            antialiased=False)
            cset = ax_res.contour(X_res, Y_res, Z_res - ZZ_res + c[0] - Z_rad, 7, zdir='z', offset=btn_limitz_res,
                            linewidths=1, cmap=cm.turbo, alpha=1)

            z_after = z_res - zz_res + c[0] - zz_rad
            ave_res = round(np.mean(z_after), 1)
            max_res = round(max(z_after))
            min_res = round(min(z_after))
            nonu_res = round(0.5*100*(max_res - min_res)/abs(ave_res), 2)
        
            ax_res.view_init(30, 240)
            ax_res.set_zlim(btn_limitz_res, top_limitz_res)
            ax_res.set_xlabel('X (mm)')
            ax_res.set_ylabel('Y (mm)')
            ax_res.set_zlabel(unit)
                 
            ax_res.set_title(
                            'After both tilting and radial pattern are corrected \n Mean = %.1f Å, NonU = %.2f pct'
                            %(ave_res, nonu_res), fontsize=6)
            fig_res.tight_layout(pad=1)    
            plot_res ="residual"
            plt.savefig(plot_res, bbox_inches='tight')
            img_res = Image.open("residual.png")
            plot_res_resized = img_res.resize((356, 356))
            plot_res_resized.save("residual_resized.png")     
            img_res = ImageTk.PhotoImage(Image.open("residual_resized.png"))
            self.canvas_graph.create_image(1088, 0, anchor="nw", image=img_res)
            os.remove("residual_resized.png")
            plt.clf()
            plt.close(fig_res)
        else:
            pass
        
        if (check_box2 == 1):
            os.remove("residual.png")
            os.remove("radius.png")
            os.remove("slope2.png")
            os.remove("slope1.png")
            os.remove("cdf.png")
            os.remove("hist.png")
            os.remove("scatter2.png")
            os.remove("scatter1.png")
            os.remove("3d.png")
            os.remove("cross_section2.png")
            os.remove("cross_section1.png")
            os.remove("contour.png")
        else:
            os.remove("3d.png")
            os.remove("cross_section2.png")
            os.remove("cross_section1.png")
            os.remove("contour.png")

        tk.mainloop()

# for saving graphs
    def save_file(self):
        global id
        global df
        global unit
        global N1
        
        self.canvas_graph.delete('all')       
        id = self.entry_id.get()
        ID = id
        unit = self.entry_unit.get()

        df5['X'] = df5['X'].apply(float)
        df5['Y'] = df5['Y'].apply(float)

        value1 = self.variable_cal.get()
        if (value1 == 'A-B'):
            df5['A'] = df5['A'].apply(float)
            df5['B'] = df5['B'].apply(float)
            df5["Z"] = df5["A"] - df5["B"]
            z = df5["Z"]
            data = [df5["X"], df5["Y"], z]
            headers = ["A", "B", "Z"]
            orignal_df = pd.concat(data, axis=1, keys=headers)
            
        elif (value1 == '(A-B)/t'):
            df5['A'] = df5['A'].apply(float)
            df5['B'] = df5['B'].apply(float)
            df5["Z"] = df5["A"] - df5["B"]
            z = df5["Z"]
            data = [df5["X"], df5["Y"], z]
            headers = ["A", "B", "Z"]
            orignal_df = pd.concat(data, axis=1, keys=headers)
            
        elif (value1 == 'B-A'):
            df5['A'] = df5['A'].apply(float)
            df5['B'] = df5['B'].apply(float)
            df5["Z"] = - df5["A"] + df5["B"]
            z = df5["Z"]
            data = [df5["X"], df5["Y"], z]
            headers = ["A", "B", "Z"]
            orignal_df = pd.concat(data, axis=1, keys=headers)
            
        elif (value1 == '(B-A)/t'):
            df5['A'] = df5['A'].apply(float)
            df5['B'] = df5['B'].apply(float)
            df5["Z"] = - df5["A"] + df5["B"]
            z = df5["Z"]
            data = [df5["X"], df5["Y"], z]
            headers = ["A", "B", "Z"]
            orignal_df = pd.concat(data, axis=1, keys=headers)
            
        elif (value1 == 'A'):
            df5['A'] = df5['A'].apply(float)
            df5['B'] = 0
            df5['B'] = df5['B'].apply(float)
            df5["Z"] = df5["A"] - df5["B"]
            z = df5["Z"]
            data = [df5["X"], df5["Y"], z]
            headers = ["A", "B", "Z"]
            orignal_df = pd.concat(data, axis=1, keys=headers)
        else:
            pass

        value = self.radio_sigma.get()
        if (value == 1):
            sigma_var = float(self.entry_sigma.get())
            z_scores = np.abs(stats.zscore(orignal_df))
            df = orignal_df[(z_scores < sigma_var).all(axis=1)]
            N1 = str(orignal_df.shape[0] - df.shape[0])
        else:
            N2 = int(self.entry_outlier.get())
            orignal_df["dist"] = np.abs(orignal_df["Z"]  - np.mean(orignal_df["Z"]))
            sort_df = orignal_df.sort_values(by="dist", ascending=False)
            df = sort_df.iloc[N2:]
            Z = df["Z"]
            stdev = np.std(Z)
            top = max(df["dist"])
            sigma2 = round(top/stdev, 2)

# create folder and excel
        import xlsxwriter
        from datetime import datetime
        import time
        
        now = datetime.now()
        dt_string = now.strftime("%m-%d-%Y %H-%M-%S")      
        check_dir = os.path.isdir('./Saved graphs')
        if (check_dir == FALSE):
            os.makedirs('Saved Graphs')
        else:
            pass
        save_path = './Saved graphs'
        file_name = id + "  " + dt_string + ".xlsx"
        complete_name = os. path. join(save_path, file_name)
        workbook = xlsxwriter.Workbook(complete_name)
        worksheet = workbook.add_worksheet()
        worksheet.set_column('A:A', 10)
        worksheet.set_column('B:B', 10)
        worksheet.set_column('C:C', 10)
        worksheet.set_column('D:D', 10)
        worksheet.set_column('E:E', 10)        

# 2d contour
        fig1 = plt.figure(figsize=(2.5, 2.5))
        spec1 = gridspec.GridSpec(ncols=1, nrows=1)
        ax1 = fig1.add_subplot(spec1[0, 0])
        
        degree_theta = float(self.rotation_entry.get())
        radian_theta = math.radians(degree_theta)
        if (degree_theta == 0):
            x2d = df["A"]
            y2d = df["B"]
        else:
            x_original = df["A"]
            y_original = df["B"]        
            x2d = math.cos(radian_theta)*x_original + math.sin(radian_theta)*y_original
            y2d = - math.sin(radian_theta)*x_original + math.cos(radian_theta)*y_original
        z2d = df["Z"]
        
        x_grid = np.linspace(np.min(x2d), np.max(x2d), 300)
        y_grid = np.linspace(np.min(y2d), np.max(y2d), 300)
        X2d, Y2d = np.meshgrid(x_grid, y_grid)
        Z2d = griddata((x2d, y2d), z2d, (X2d, Y2d), method="cubic")
        
        Ave = round(np.mean(z2d), 1)
        Std = np.std(z2d)
        Std_percent = round(100*Std/abs(Ave), 2)
        Max = round(max(z2d), 1)
        Min = round(min(z2d), 1)
        Nonu = round(0.5*100*(Max-Min)/abs(Ave), 2)
        Range = round(max(z2d) - min(z2d), 1)
        
        value1 = self.variable_cal.get()
        if (value1 == '(A-B)/t'):
            time_pre = self.entry_run_time.get()
            time = float(time_pre)
            Rate_pre = abs(Ave)*(60/time)
            Rate = round(Rate_pre, 1)         
            plt.text(-75,-330, 'Rate (' + unit + '/min)')
            plt.text(45,-330, Rate)
            worksheet.write('B4', Rate)
            worksheet.write('A9', 'Rate (' + unit + '/min)')
            worksheet.write('B9', Rate)
        elif (value1 == '(B-A)/t'):
            time_pre = self.entry_run_time.get()
            time = float(time_pre)
            Rate_pre = abs(Ave)*(60/time)
            Rate = round(Rate_pre, 1)         
            plt.text(-75,-330, 'Rate (' + unit + '/min)')
            plt.text(45,-330, Rate)
            worksheet.write('A9', 'Rate (' + unit + '/min)')
            worksheet.write('B9', Rate)
        else:
            pass

        check_box = self.var_limits.get() 
        if (check_box == 1):
            limits_bt_pre = self.limits_entry.get()
            limits_bt = abs(float(limits_bt_pre))
            limits_b = Ave - limits_bt/2
            limits_t = Ave + limits_bt/2
            num = int(self.entry_contour.get())
            if (num >100):
                num = 100
                levels = np.linspace(limits_b, limits_t, num)
            else:
                levels = np.linspace(limits_b, limits_t, num)
            cp = plt.contourf(X2d, Y2d, Z2d, levels = levels, cmap=plt.cm.turbo, alpha=0.95)
            fig1.colorbar(cp, ax=ax1, shrink=0.5, orientation="vertical")
        else:
            contour = int(self.entry_contour.get()) 
            if (contour >100):
                contour = 100
            else:
                pass
            cp = plt.contourf(X2d, Y2d, Z2d, contour, cmap=plt.cm.turbo, alpha=0.95)
            cbar1 = fig1.colorbar(cp, ax=ax1, shrink=0.5, orientation="vertical")
            cbar1.ax.locator_params(nbins=6)
        
        plt.text(-75,-240, 'Mean (' + unit + ')')
        plt.text(-75,-255, 'NonU (%)')
        plt.text(-75,-270, 'StdD (%)')
        plt.text(-75,-285, 'Max (' + unit + ')')
        plt.text(-75,-300, 'Min (' + unit + ')')
        plt.text(-75,-315, 'Range (' + unit + ')')
        plt.text(45,-240, Ave)
        plt.text(45,-255, Nonu)                                      
        plt.text(45,-270, Std_percent)
        plt.text(45,-285, Max)
        plt.text(45,-300, Min)
        plt.text(45,-315, Range)

        worksheet.write('A1', 'Time')
        worksheet.write('A2', 'ID')
        worksheet.write('A3', 'Mean (' + unit + ')')
        worksheet.write('A4', 'NonU (%)')
        worksheet.write('A5', 'StdD (%)')
        worksheet.write('A6', 'Max (' + unit + ')')
        worksheet.write('A7', 'Min (' + unit + ')')
        worksheet.write('A8', 'Range (' + unit + ')')
        worksheet.write('B1', dt_string)
        worksheet.write('B2', id)
        worksheet.write('B3', abs(Ave))
        worksheet.write('B4', Nonu)
        worksheet.write('B5', Std_percent)
        worksheet.write('B6', Max)
        worksheet.write('B7', Min)
        worksheet.write('B8', Range)
        worksheet.write('C2', film1)
        
        worksheet.write('A11', 'X')
        worksheet.write('B11', 'Y')
        worksheet.write('C11', 'A')
        worksheet.write('D11', 'B')
        worksheet.write('E11', 'Z ')
        worksheet.write_column(11, 0, x2d)
        worksheet.write_column(11, 1, y2d)
        worksheet.write_column(11, 4, z)
                        
        if (value1 == 'A'):
            worksheet.write_column(11, 2, df5['A'])
        elif (value1 == 'B'):
            worksheet.write_column(11, 3, df5['B'])
        else:
            worksheet.write_column(11, 2, df5['A'])
            worksheet.write_column(11, 3, df5['B'])
            
        value = self.radio_sigma.get()
        if (value == 1):
            plt.text(-75,-345, 'Points removed')
            plt.text(45,-345, N1)
        else:
            plt.text(-75,-345, 'Sigma:')
            plt.text(45,-345, sigma2)
        
        circ = Circle((0, 0), 150, facecolor='None', edgecolor='black', lw=0.2, alpha=0.4)
        ax1.add_patch(circ)
        ax1.set_aspect('equal', adjustable='box')

        var3 = self.variable3.get()
        if (var3 == "Value"):
            for x2d, y2d, z2d in zip(x2d, y2d, z2d):
                label = "{:.0f}".format(z2d)
                plt.annotate(label, (x2d, y2d), textcoords="offset points", xytext=(0, -3), ha='center', 
                            fontsize=4.3, alpha=0.85)
                ax1.scatter(x2d, y2d, marker='o', s=1, color='k', alpha=0.0001)
        elif (var3 == "Dot"):
            for x2d, y2d, z2d in zip(x2d, y2d, z2d):
                ax1.scatter(x2d, y2d, marker='o', s=1, color='k', alpha=0.3)
        elif (var3 == "None"):
            for x2d, y2d, z2d in zip(x2d, y2d, z2d):
                label = "{:.0f}".format(z2d)
                ax1.scatter(x2d, y2d, marker='o', s=1, color='k', alpha=0.0001)
        else:
            for x2d, y2d, z2d in zip(x2d, y2d, z2d):
                label = "{:.0f}".format(z2d)
                if (z2d > Ave):
                    ax1.scatter(x2d, y2d, marker='+', s=13, linewidths=0.3, color='k', alpha=0.7)
                else:
                    ax1.scatter(x2d, y2d, marker='_', s=13, linewidths=0.3, color='k', alpha=0.7)

        ax1.set_xlabel('X (mm)')
        ax1.set_ylabel('Y (mm)')    
        ax1.set_title(id + '    ' + film1)
    
        fig1.tight_layout(pad=0.1)
        plot_id1 = "contour"
        plt.savefig(plot_id1, bbox_inches='tight')        
        img1 = Image.open("contour.png")
        plot_id1_resized = img1.resize((350, 380))
        plot_id1_resized.save("contour_resized.png")     
        img11 = ImageTk.PhotoImage(Image.open("contour_resized.png"))
        self.canvas_graph.create_image(20, 35, anchor="nw", image=img11)
        worksheet.insert_image('G2', 'contour.png', {'x_scale': 1.02, 'y_scale': 1.02})
        os.remove("contour_resized.png")
        plt.clf()
        plt.close(fig1)
            
# cross_section 1
        fig_crs1 = plt.figure(figsize=(2.3, 1.8))     
        spec_crs1 = gridspec.GridSpec(ncols=1, nrows=1)
        ax_crs1 = fig_crs1.add_subplot(spec_crs1[0, 0])
        ax_crs1.set_xlabel('Cross section (mm)')
        ax_crs1.set_ylabel(unit)
        ax_crs1.set(title=id)

        degree_theta = float(self.rotation_entry.get())
        radian_theta = math.radians(degree_theta)
        if (degree_theta == 0):
            x_crs1 = df["A"]
            y_crs1 = df["B"]
        else:
            x_oric1 = df["A"]
            y_oric1 = df["B"]
            x_crs1 = math.cos(radian_theta)*x_oric1+ math.sin(radian_theta)*y_oric1
            y_crs1 = - math.sin(radian_theta)*x_oric1 + math.cos(radian_theta)*y_oric1
        z_crs1 = df["Z"]

        n_crs1 = 80
        x_grid_crs1 = np.linspace(np.min(x_crs1), np.max(x_crs1), n_crs1)
        y_grid_crs1 = np.linspace(np.min(y_crs1), np.max(y_crs1), n_crs1)
        X_crs1, Y_crs1 = np.meshgrid(x_grid_crs1, y_grid_crs1)
        
        Y_crs1 = 0
        Z_crs1 = griddata((x_crs1, y_crs1), z_crs1, (X_crs1, Y_crs1), method="cubic")

        degree_crs1_90 = 90
        radian_crs1_90 = math.radians(degree_crs1_90)     
        x_crs1_90 = math.cos(radian_crs1_90)*x_crs1 + math.sin(radian_crs1_90)*y_crs1
        y_crs1_90 = - math.sin(radian_crs1_90)*x_crs1 + math.cos(radian_crs1_90)*y_crs1

        Y_crs1 = 0
        Z_crs1_90 = griddata((x_crs1_90, y_crs1_90), z_crs1, (X_crs1, Y_crs1), method="cubic")

        ave1 = np.mean(z_crs1)
        count_row = n_crs1
        Std = np.std(z_crs1)
        sigma_var = float(self.entry_sigma.get())

        Z_crs1_1 = [ave1 - 3*Std]*count_row 
        Z_crs1_2 = [ave1]*count_row 
        Z_crs1_3 = [ave1 + 3*Std]*count_row

        ax_crs1.plot(X_crs1, Z_crs1_1, '.', color='grey', markersize=0.015, alpha=0.165)
        ax_crs1.plot(X_crs1, Z_crs1_2, '.', color='grey', markersize=0.015, alpha=0.165)
        ax_crs1.plot(X_crs1, Z_crs1_3, '.', color='grey', markersize=0.015, alpha=0.165)

        ax_crs1.plot(X_crs1, Z_crs1, 'o', color='red', markersize=0.01)
        ax_crs1.plot(X_crs1, Z_crs1_90, 'o', color='blue', markersize=0.01)
        
        ax_crs1.legend(("Red: X-X", "Blue: Y-Y"), frameon=False, loc='upper right')
        top_x_crs1 = max(x_crs1)
        btn_x_crs1 = min(x_crs1)
        ax_crs1.set_xlim(btn_x_crs1, top_x_crs1)

        check_box = self.var_limits.get() 
        if (check_box == 1):
            limits_y_pre = self.limits_entry.get()
            limits_y = abs(float(limits_y_pre))
            btn_y_crs1 = Ave - limits_y/2
            top_y_crs1 = Ave + limits_y/2
            ax_crs1.set_ylim(btn_y_crs1, top_y_crs1)            
        else:
            top_y_crs1 = Ave + Range*1.8
            btn_y_crs1 = Ave - Range*1.8   
            ax_crs1.set_ylim(btn_y_crs1, top_y_crs1)

        fig_crs1.tight_layout(pad=1)    
        plot_id_crs1 ="cross_section1"
        plt.savefig(plot_id_crs1, bbox_inches='tight')
        img_crs1 = Image.open("cross_section1.png")
        plot_crs1_resized = img_crs1.resize((347, 247))
        plot_crs1_resized.save("cross_resized1.png")     
        img_crs1 = ImageTk.PhotoImage(Image.open("cross_resized1.png"))
        self.canvas_graph.create_image(0, 420, anchor="nw", image=img_crs1)
        worksheet.insert_image('G17', 'cross_section1.png', {'x_scale': 1.12, 'y_scale': 1.12})
        os.remove("cross_resized1.png")
        plt.clf()
        plt.close(fig_crs1)

# cross_section 2
        fig_crs2 = plt.figure(figsize=(2.3, 1.8))     
        spec_crs2 = gridspec.GridSpec(ncols=1, nrows=1)
        ax_crs2 = fig_crs2.add_subplot(spec_crs2[0, 0])
        ax_crs2.set_xlabel('Cross section (mm)')
        ax_crs2.set_ylabel(unit)
        ax_crs2.set(title=id)

        degree_theta = float(self.rotation_entry.get())
        radian_theta = math.radians(degree_theta)
        if (degree_theta == 0):
            x_crs2 = df["A"]
            y_crs2 = df["B"]
        else:
            x_oric2 = df["A"]
            y_oric2 = df["B"]
            x_crs2 = math.cos(radian_theta)*x_oric2+ math.sin(radian_theta)*y_oric2
            y_crs2 = - math.sin(radian_theta)*x_oric2 + math.cos(radian_theta)*y_oric2
        z_crs2 = df["Z"]

        x_grid_crs2 = np.linspace(np.min(x_crs2), np.max(x_crs2), 80)
        y_grid_crs2 = np.linspace(np.min(y_crs2), np.max(y_crs2), 80)
        X_crs2, Y_crs2 = np.meshgrid(x_grid_crs2, y_grid_crs2)

        Y_max = max(y_crs2)
        Y_min = min(y_crs2)
        ran_crs2 = (max(y_crs2) - min(y_crs2))/4
        Y_cros1 = - ran_crs2
        Y_cros3 = 0
        Y_cros5 = ran_crs2

        Z_crs2_1 = griddata((x_crs2, y_crs2), z_crs2, (X_crs2, Y_cros1), method="cubic")
        Z_crs2_3 = griddata((x_crs2, y_crs2), z_crs2, (X_crs2, Y_cros3), method="cubic")
        Z_crs2_5 = griddata((x_crs2, y_crs2), z_crs2, (X_crs2, Y_cros5), method="cubic")

        ax_crs2.plot(X_crs2, Z_crs1_1, '.', color='grey', markersize=0.015, alpha=0.165)
        ax_crs2.plot(X_crs2, Z_crs1_2, '.', color='grey', markersize=0.015, alpha=0.165)
        ax_crs2.plot(X_crs2, Z_crs1_3, '.', color='grey', markersize=0.015, alpha=0.165)
        
        ax_crs2.plot(X_crs2, Z_crs2_5, 'o', color='orange', markersize=0.01)      
        ax_crs2.plot(X_crs2, Z_crs2_3, 'o', color='red', markersize=0.01)
        ax_crs2.plot(X_crs2, Z_crs2_1, 'o', color='green', markersize=0.01)
           
        ax_crs2.legend(("Green: 1st X-X", "Red: 2nd X-X", "Yellow: 3rd X-X"), frameon=False,
                           loc='upper right')
        top_x_crs2 = max(x_crs2)
        btn_x_crs2 = min(x_crs2)
        ax_crs2.set_xlim(btn_x_crs2, top_x_crs2)

        check_box = self.var_limits.get() 
        if (check_box == 1):
            limits_y_pre = self.limits_entry.get()
            limits_y = abs(float(limits_y_pre))
            btn_y_crs2 = Ave - limits_y/2
            top_y_crs2 = Ave + limits_y/2
            ax_crs2.set_ylim(btn_y_crs2, top_y_crs2)            
        else:
            top_y_crs2 = Ave + Range*1.8
            btn_y_crs2 = Ave - Range*1.8   
            ax_crs2.set_ylim(btn_y_crs2, top_y_crs2)

        fig_crs2.tight_layout(pad=1)    
        plot_id_crs2 ="cross_section2"
        plt.savefig(plot_id_crs2, bbox_inches='tight')
        img_crs2 = Image.open("cross_section2.png")
        plot_crs2_resized = img_crs2.resize((347, 247))
        plot_crs2_resized.save("cross_resized2.png")     
        img_crs2 = ImageTk.PhotoImage(Image.open("cross_resized2.png"))
        self.canvas_graph.create_image(370, 420, anchor="nw", image=img_crs2)
        worksheet.insert_image('L17', 'cross_section2.png', {'x_scale': 1.12, 'y_scale': 1.12})
        os.remove("cross_resized2.png")
        plt.clf()
        plt.close(fig_crs2)

# 3d_contour
        fig3d = plt.figure(figsize=(2.5, 2.5))
        ax3d = plt.axes(projection="3d")
        ax3d.view_init(30, 240)
        degree_theta = float(self.rotation_entry.get())
        radian_theta = math.radians(degree_theta)
        if (degree_theta == 0):
            x3d = df["A"]
            y3d = df["B"]
        else:
            degree_theta = float(self.rotation_entry.get())
            radian_theta = math.radians(degree_theta)
            x_orig = df["A"]
            y_orig = df["B"]
            x3d = math.cos(radian_theta)*x_orig + math.sin(radian_theta)*y_orig
            y3d = - math.sin(radian_theta)*x_orig + math.cos(radian_theta)*y_orig
            
        z3d = df["Z"] 

        x3d_grid = np.linspace(np.min(x3d), np.max(x3d), 120)
        y3d_grid = np.linspace(np.min(y3d), np.max(y3d), 120)
        X3d, Y3d = np.meshgrid(x3d_grid, y3d_grid)
        
        Z3d = griddata((x3d, y3d), z3d, (X3d, Y3d), method="cubic")
        ax3d.contour3D(X3d, Y3d, Z3d, 750, cmap='turbo', alpha=1, antialiased=False)
 
        check_box = self.var_limits.get() 
        if (check_box == 1):
            limits_z_pre = self.limits_entry.get()
            limits_z = abs(float(limits_z_pre))
            btn_limitz3d = np.mean(z3d) - limits_z/2
            top_limitz3d = np.mean(z3d) + limits_z/2
            ax3d.set_zlim(btn_limitz3d, top_limitz3d)          
        else:
            top_limitz3d = np.mean(z3d) + (max(z3d)-min(z3d))*1.5
            btn_limitz3d = np.mean(z3d) - (max(z3d)-min(z3d))*4.5
            ax3d.set_zlim(btn_limitz3d, top_limitz3d)

        cset = ax3d.contour(X3d, Y3d, Z3d, 10, zdir='z', offset=btn_limitz3d, linewidths=1,
                            cmap=cm.turbo, alpha=0.9)
            
        ax3d.set_xlabel('X (mm)')
        ax3d.set_ylabel('Y (mm)')
        ax3d.set_zlabel(unit)
        ax3d.set_title('Original 3d map \n Mean = %.1f Å, NonU = %.2f pct' %(Ave, Nonu), fontsize=6)
        
        fig3d.tight_layout(pad=1)    
        plot_id3d ="3d"
        plt.savefig(plot_id3d, bbox_inches='tight')
        img3d = Image.open("3d.png")
        plot_id3d_resized = img3d.resize((356, 356))
        plot_id3d_resized.save("3d_resized.png")     
        img33d = ImageTk.PhotoImage(Image.open("3d_resized.png"))
        self.canvas_graph.create_image(375, 0, anchor="nw", image=img33d)
        worksheet.insert_image('L1', '3d.png', {'x_scale': 1.085, 'y_scale': 1.085})
        os.remove("3d_resized.png")
        plt.clf()
        plt.close(fig3d)

# 3d animated
        from matplotlib.animation import FuncAnimation

        check_box3 = self.var_animated.get() 
        if (check_box3 == 1):
            fig_3d_rot = plt.figure()
            fig_3d_rot.set_size_inches(2, 2)
            ax_3d_rot = plt.axes(projection='3d')
            ax_3d_rot.set_facecolor((1, 1, 1))

            check_box = self.var_limits.get() 
            if (check_box == 1):
                limits_z_pre = self.limits_entry.get()
                top_limitz3d_animated = float(limits_z_pre)
                limits_z_pre2 = self.limits_entry.get()
                btn_limitz3d_animated = float(limits_z_pre2)
                ax_3d_rot.set_zlim(btn_limitz3d, top_limitz3d)          
            else:
                top_limitz3d_animated = np.mean(z3d) + (max(z3d)-min(z3d))*1.85
                btn_limitz3d_animated = np.mean(z3d) - (max(z3d)-min(z3d))*1.85
                ax_3d_rot.set_zlim(btn_limitz3d_animated, top_limitz3d_animated)

            ax_3d_rot.set_xlabel('X', fontsize=4)
            ax_3d_rot.set_ylabel('Y', fontsize=4)
            ax_3d_rot.set_zlabel('Z', fontsize=4)
            ax_3d_rot.set_title(id, fontsize=5)

            ax_3d_rot.plot_surface(X3d, Y3d, Z3d, rstride=6, cstride=6, cmap='viridis', alpha=0.85,
                               edgecolor='none')
            plt.axis('on')

            def update(i, fig_3d_rot, ax_3d_rot):
                ax_3d_rot.view_init(elev=30., azim=i)
                return fig_3d_rot, ax_3d_rot

            anim = FuncAnimation(fig_3d_rot, update, frames=np.arange(0, 360, 1), repeat=True,
                                fargs=(fig_3d_rot, ax_3d_rot), interval=100)
            anim.save('rgb_cube.gif', dpi=60, writer='pillow')
            plt.pause(0.1)
            os.remove('rgb_cube.gif')
        else:
            pass
            
# plot all
        check_box2 = self.var_decom.get() 
        if (check_box2 == 1):
            
# scatter1
            fig5 = plt.figure(figsize=(2.3, 1.7))
            spec5 = gridspec.GridSpec(ncols=1, nrows=1)

            df["radius (mm)"] = np.sqrt(df["A"]**2 + df["B"]**2)
            df["theta_pre (degree)"] = np.arctan2(df["B"], df["A"])*180/np.pi
            df["theta (degree)"] = df["theta_pre (degree)"].apply(lambda x: x + 360 if x < 0 else x)

            ax5 = fig5.add_subplot(spec5[0, 0])       
            a1 = 1
            b1 = 0.1
            df["u"] = a1*df["radius (mm)"] + b1*df["theta (degree)"]
            df = df.sort_values(by="u")
            df["Point"] = range(1, 1 + len(df["u"]))
            x1 = df["Point"] 
            y1 = df["Z"]

            ave1 = np.mean(y1)
            count_row = df.shape[0]
            Std = np.std(y1)
            sigma_var = float(self.entry_sigma.get())

            y2 = [ave1 - 3*Std]*count_row 
            y3 = [ave1]*count_row 
            y4 = [ave1 + 3*Std]*count_row

            ax5.plot(x1, y1, 'o', markersize=1.5, alpha=0.7)
            ax5.plot(x1, y2, '-', markersize=0.001, c='grey', alpha=0.15)
            ax5.plot(x1, y3, '-', markersize=0.001, c='grey', alpha=0.15)
            ax5.plot(x1, y4, '-', markersize=0.001, c='grey', alpha=0.15)

            ax5.set_xlabel('Measurement point', fontsize=5)
            ax5.set_ylabel(unit, fontsize=5)
            ax5.set(title = 'As radius increasing')

            check_box = self.var_limits.get() 
            if (check_box == 1):
                limit1_bt_pre = self.limits_entry.get()
                limit1_bt = abs(float(limit1_bt_pre))
                btn_limit1 = Ave - limit1_bt/2
                top_limit1 = Ave + limit1_bt/2
                ax5.set_ylim(btn_limit1, top_limit1)
            else:
                top_limit1 = Ave + Range*1.8
                btn_limit1 = Ave - Range*1.8
                ax5.set_ylim(btn_limit1, top_limit1)

            fig5.tight_layout(pad=1)
            plot_id5 ="scatter1"
            plt.savefig(plot_id5, bbox_inches='tight')
            img5 = Image.open("scatter1.png")
            plot_id5_resized = img5.resize((352, 247))
            plot_id5_resized.save("scatter1_resized.png")      
            img5 = ImageTk.PhotoImage(Image.open("scatter1_resized.png")) 
            self.canvas_graph.create_image(0, 690, anchor="nw", image=img5)
            worksheet.insert_image('G28', 'scatter1.png', {'x_scale': 1.19, 'y_scale': 1.19})
            os.remove("scatter1_resized.png")
            plt.clf()
            plt.close(fig5)

# scatter2
            fig6 = plt.figure(figsize=(2.3, 1.7))
            spec6 = gridspec.GridSpec(ncols=1, nrows=1)

            df["radius (mm)"] = np.sqrt(df["A"]**2 + df["B"]**2)
            df["theta_pre (degree)"] = np.arctan2(df["B"], df["A"])*180/np.pi
            df["theta (degree)"] = df["theta_pre (degree)"].apply(lambda x: x + 360 if x < 0 else x)

            ax6 = fig6.add_subplot(spec6[0, 0]) 
            a2 = 0.1
            b2 = 1
            df["v"] = a2*df["radius (mm)"] + b2*df["theta (degree)"]    
            df = df.sort_values(by="v")
            df["Point"] = range(1, 1 + len(df["v"]))
            x2 = df["Point"] 
            y2 = df["Z"] 

            ave1 = np.mean(y1)
            count_row = df.shape[0]
            Std = np.std(y1)
            sigma_var = float(self.entry_sigma.get())

            y2_2 = [ave1 - 3*Std]*count_row 
            y2_3 = [ave1]*count_row 
            y2_4 = [ave1 + 3*Std]*count_row 

            ax6.plot(x2, y2, 'o', markersize=1.5, c='red', alpha=0.7)
            ax6.plot(x2, y2_2, '-', markersize=0.001, c='grey', alpha=0.15)
            ax6.plot(x2, y2_3, '-', markersize=0.001, c='grey', alpha=0.15)
            ax6.plot(x2, y2_4, '-', markersize=0.001, c='grey', alpha=0.15)

            ax6.set_xlabel('Measurement point', fontsize=5)
            ax6.set_ylabel(unit, fontsize=5)
            ax6.set(title='As theta increasing')

            check_box = self.var_limits.get() 
            if (check_box == 1):
                limit2_bt_pre = self.limits_entry.get()
                limit2_bt = abs(float(limit2_bt_pre))
                btn_limit2 = Ave - limit2_bt/2
                top_limit2 = Ave + limit2_bt/2
                ax6.set_ylim(btn_limit2, top_limit2)
            else:
                top_limit2 = Ave + Range*1.8
                btn_limit2 = Ave - Range*1.8
                ax6.set_ylim(btn_limit2, top_limit2)

            fig6.tight_layout(pad=1)
            plot_id6 ="scatter2"
            plt.savefig(plot_id6, bbox_inches='tight')
            img6 = Image.open("scatter2.png")
            plot_id6_resized = img6.resize((352, 247))
            plot_id6_resized.save("scatter2_resized.png")      
            img6 = ImageTk.PhotoImage(Image.open("scatter2_resized.png")) 
            self.canvas_graph.create_image(370, 690, anchor="nw", image=img6)
            worksheet.insert_image('L28', 'scatter2.png', {'x_scale': 1.19, 'y_scale': 1.19})
            os.remove("scatter2_resized.png")
            plt.clf()
            plt.close(fig6)

# histogram
            fig10 = plt.figure(figsize=(2.5, 1.9))
            spec10 = gridspec.GridSpec(ncols=1, nrows=1)       
            ax10 = fig10.add_subplot(spec10[0, 0])
            ax10.set_xlabel(unit, fontsize=6)
            ax10.set_ylabel('Counts', fontsize=6)
            ax10.set(title = 'Histogram')
            z_hist = df["Z"]
            bins = round(len(z_hist)/5)

            z1 = ave1 - 3*Std
            z2 = ave1
            z3 = ave1 + 3*Std

            ax10.hist(z_hist, bins, color='darkblue', alpha=0.5)
            plt.axvline(z1, color='grey', linestyle=':', linewidth=0.5)
            plt.axvline(z2, color='grey', linestyle=':', linewidth=0.5)
            plt.axvline(z3, color='grey', linestyle=':', linewidth=0.5)

            check_box = self.var_limits.get() 
            if (check_box == 1):
                limit_pre = self.limits_entry.get()
                limit_abs = abs(float(limit_pre))
                btn_limit = Ave - limit_abs/2
                top_limit = Ave + limit_abs/2
                ax10.set_xlim((btn_limit, top_limit))
            else:
                btn_limit = Ave - Range*1.3
                top_limit = Ave + Range*1.3
                ax10.set_xlim((btn_limit, top_limit))

            fig10.tight_layout(pad=1)
            plot_id10 ="hist"
            plt.savefig(plot_id10, bbox_inches='tight')
            img10 = Image.open("hist.png")
            plot_id10_resized = img10.resize((316, 242))
            plot_id10_resized.save("hist_resized.png")      
            img10 = ImageTk.PhotoImage(Image.open("hist_resized.png")) 
            self.canvas_graph.create_image(35, 955, anchor="nw", image=img10)
            worksheet.insert_image('G39', 'hist.png', {'x_scale': 1.05, 'y_scale': 1.05})
            os.remove("hist_resized.png")
            plt.clf()
            plt.close(fig10)

# cumulative distribution
            fig_cdf = plt.figure(figsize=(2.5, 1.9))
            spec_cdf = gridspec.GridSpec(ncols=1, nrows=1)       
            ax_cdf = fig_cdf.add_subplot(spec_cdf[0, 0])
            ax_cdf.set_xlabel(unit, fontsize=6)
            ax_cdf.set_ylabel('Percent', fontsize=6)
            ax_cdf.set(title = 'Cumulative distribution')

            x_cdf_pre = df["Z"]
            x_cdf = np.sort(x_cdf_pre)
            y_cdf = 100*np.arange(len(x_cdf))/float(len(x_cdf))

            ave = np.mean(x_cdf_pre)
            count_row = df.shape[0]
            Std = np.std(x_cdf_pre)
            sigma_var = float(self.entry_sigma.get())

            x_cdf_2= [ave - 3*Std]*count_row 
            x_cdf_3 = [ave]*count_row 
            x_cdf_4 = [ave + 3*Std]*count_row 

            ax_cdf.plot(x_cdf, y_cdf, 'o', markersize=1.8, color='m', alpha=0.7)
            ax_cdf.plot(x_cdf_2, y_cdf, '-', markersize=0.001, c='grey', alpha=0.15)
            ax_cdf.plot(x_cdf_3, y_cdf, '-', markersize=0.001, c='grey', alpha=0.15)
            ax_cdf.plot(x_cdf_4, y_cdf, '-', markersize=0.001, c='grey', alpha=0.15)

            check_box = self.var_limits.get() 
            if (check_box == 1):
                limit_pre = self.limits_entry.get()
                limit_abs = abs(float(limit_pre))
                btn_limit = Ave - limit_abs/2
                top_limit = Ave + limit_abs/2
                ax_cdf.set_xlim((btn_limit, top_limit))
            else:
                top_cdf = Ave + Range*1.3
                btn_cdf = Ave - Range*1.3
                ax_cdf.set_xlim(btn_cdf, top_cdf)

            fig_cdf.tight_layout(pad=1)
            plot_id_cdf ="cdf"
            plt.savefig(plot_id_cdf, bbox_inches='tight')
            img_cdf = Image.open("cdf.png")
            plot_id_cdf_resized = img_cdf.resize((328, 247))
            plot_id_cdf_resized.save("cdf_resized.png")      
            img_cdf = ImageTk.PhotoImage(Image.open("cdf_resized.png")) 
            self.canvas_graph.create_image(395, 955, anchor="nw", image=img_cdf)
            worksheet.insert_image('L39', 'cdf.png', {'x_scale': 1.05, 'y_scale': 1.05})
            os.remove("cdf_resized.png")
            plt.clf()
            plt.close(fig_cdf)

# decompose_tilt_corrected
            fig_slp1 = plt.figure(figsize=(2.5, 2.5))
            ax_slp1 = plt.axes(projection="3d")
            degree_theta = float(self.rotation_entry.get())
            radian_rotation = math.radians(degree_theta)
                                    
            if (degree_theta == 0):
                df["x_tc"] = df["A"]
                df["y_tc"] = df["B"]
            else:
                x_orir = df["A"]
                y_orir = df["B"]
                df["x_tc"] = math.cos(radian_rotation)*x_orir + math.sin(radian_rotation)*y_orir
                df["y_tc"] = -math.sin(radian_rotation)*x_orir + math.cos(radian_rotation)*y_orir

            x_slp1 = df["x_tc"]
            y_slp1 = df["y_tc"]
            z_slp1 = df["Z"]

            x_tmp = df[['x_tc', 'y_tc']].values.reshape(-1, 2)
            max_rows = max(len(x_slp1), len(y_slp1))
            x_ones = np.ones([max_rows, 1])
            x = np.append(x_ones, x_tmp, axis=1)
            y_tmp = z_slp1.tolist()
            y = np.array(y_tmp)

            def mse(coef, x, y):
                return np.mean((np.dot(x, coef) - y)**2)/2
            def gradients(coef, x, y):
                return np.mean(x.transpose()*(np.dot(x, coef) - y), axis=1)
            def multilinear_regression(coef, x, y, lr, b1=0.9, b2=0.999, epsilon=1e-8):
                prev_error = 0
                m_coef = np.zeros(coef.shape)
                v_coef = np.zeros(coef.shape)
                moment_m_coef = np.zeros(coef.shape)
                moment_v_coef = np.zeros(coef.shape)
                t = 0
                while True:
                    error = mse(coef, x, y)
                    if abs(error - prev_error) <= epsilon:
                        break
                    prev_error = error
                    grad = gradients(coef, x, y)
                    t += 1
                    m_coef = b1*m_coef + (1-b1)*grad
                    v_coef = b2*v_coef + (1-b2)*grad**2
                    moment_m_coef = m_coef / (1-b1**t)
                    moment_v_coef = v_coef / (1-b2**t)
                    delta = ((lr / moment_v_coef**0.5 + 1e-8)*(b1*moment_m_coef + (1-b1)*grad/(1-b1**t)))
                    coef = np.subtract(coef, delta)
                return coef
 
            coef = np.array([0, 0, 0])
            c = multilinear_regression(coef, x, y, 1e-1)
        
            check_box = self.var_limits.get() 
            if (check_box == 1):
                limits_z_pre = self.limits_entry.get()
                limits_z = abs(float(limits_z_pre))
                btn_limitz_slp1 = np.mean(z_slp1) - limits_z/2
                top_limitz_slp1 = np.mean(z_slp1) + limits_z/2
            else:
                ax_slp1.set_zlim(btn_limitz3d, top_limitz3d)
                btn_limitz_slp1 = np.mean(z_slp1) - (max(z_slp1)-min(z_slp1))*4.5
                top_limitz_slp1 = np.mean(z_slp1) + (max(z_slp1)-min(z_slp1))*1.5

            x_slp1_grid = np.linspace(np.min(x_slp1), np.max(x_slp1), 120)
            y_slp1_grid = np.linspace(np.min(y_slp1), np.max(y_slp1), 120)
            X_slp1, Y_slp1 = np.meshgrid(x_slp1_grid, y_slp1_grid)
            Z_slp1 = griddata((x_slp1, y_slp1), z_slp1, (X_slp1, Y_slp1), method="cubic")
            zz_slp1 = c[1]*x_slp1 + c[2]*y_slp1 + c[0]
            ZZ_slp1 = griddata((x_slp1, y_slp1), zz_slp1, (X_slp1, Y_slp1), method="linear")
            
            ax_slp1.contour3D(X_slp1, Y_slp1, Z_slp1 - ZZ_slp1 + c[0], 450, cmap='turbo', alpha=1,
                            antialiased=False)
            cset = ax_slp1.contour(X_slp1, Y_slp1, Z_slp1 - ZZ_slp1 + c[0], 7, zdir='z', offset=btn_limitz_slp1,
                            linewidths=1, cmap=cm.turbo, alpha=0.9)

            z_after = z_slp1 - zz_slp1 + c[0]
            ave_slp1 = round(np.mean(z_after), 1)
            max_slp1 = round(max(z_after))
            min_slp1 = round(min(z_after))
            nonu_slp1 = round(0.5*100*(max_slp1 - min_slp1)/abs(ave_slp1), 2)

            ax_slp1.set_title('After tilting is corrected \n Mean = %.1f Å, NonU = %.2f pct' %(ave_slp1,
                            nonu_slp1), fontsize=6)
            ax_slp1.set_zlim(btn_limitz_slp1, top_limitz_slp1)
            ax_slp1.view_init(30, 240)
            ax_slp1.set_xlabel('X (mm)')
            ax_slp1.set_ylabel('Y (mm)')
            ax_slp1.set_zlabel(unit)

            plot_slp1 ="slope1"
            plt.savefig(plot_slp1, bbox_inches='tight')
            img_slp1 = Image.open("slope1.png")
            plot_slp1_resized = img_slp1.resize((356, 356))
            plot_slp1_resized.save("slope1_resized.png")     
            img_slp1 = ImageTk.PhotoImage(Image.open("slope1_resized.png"))
            self.canvas_graph.create_image(732, 0, anchor="nw", image=img_slp1)
            worksheet.insert_image('Q1', 'slope1.png', {'x_scale': 1.15, 'y_scale': 1.15})
            os.remove("slope1_resized.png")
            plt.clf()
            plt.close(fig_slp1)
        
# decompose_slope
            fig_slp2 = plt.figure(figsize=(2.5, 2.5))
            ax_slp2 = plt.axes(projection="3d")
            degree_theta = float(self.rotation_entry.get())
            radian_rotation = math.radians(degree_theta)
            
            if (degree_theta == 0):
                x_slp2 = df["A"]
                y_slp2 = df["B"]
            else:
                x_orir = df["A"]
                y_orir = df["B"]
                x_slp2 = math.cos(radian_rotation)*x_orir + math.sin(radian_rotation)*y_orir
                y_slp2 = -math.sin(radian_rotation)*x_orir + math.cos(radian_rotation)*y_orir

            z_slp2 = df["Z"]
            zz_slp2 = c[1]*x_slp2 + c[2]*y_slp2 + c[0]

            actual = y_tmp
            predict = zz_slp2
            correlation_matrix = np.corrcoef(actual, predict)
            corr = correlation_matrix[0, 1]
            R_sq1 = corr**2
            
            v1 = (c[1], c[2], -10**(-7))
            v2 = (0, 0, 10**(-7))
            
            def unit_vector(vector):
                return vector / np.linalg.norm(vector)
            def angle_between(v1, v2):
                v1_u = unit_vector(v1)
                v2_u = unit_vector(v2)
                return np.arccos(np.clip(np.dot(v1_u, v2_u), -1.0, 1.0))
            angle_deg = np.round(180 - math.degrees(angle_between(v1, v2)), decimals=9, out=None)
            angle_deg2 = np.round(90 - angle_deg, decimals=9, out=None) 

            angle_skew = math.degrees(math.atan2(-c[1]/c[2], 1))
            non_acute = angle_skew + 90

            x_slp2_grid = np.linspace(np.min(x_slp2), np.max(x_slp2), 120)
            y_slp2_grid = np.linspace(np.min(y_slp2), np.max(y_slp2), 120)
            X_slp2, Y_slp2 = np.meshgrid(x_slp2_grid, y_slp2_grid)
            Z_slp2 = griddata((x_slp2, y_slp2), zz_slp2, (X_slp2, Y_slp2), method="linear")
                
            btn_limitz_slp2 = btn_limitz3d
            top_limitz_slp2 = top_limitz3d
            Z_slp3 = Z_slp2 - c[0] + np.mean(z3d)
            
            ax_slp2.contour3D(X_slp2, Y_slp2, Z_slp3, 150, cmap='turbo', alpha=0.35, antialiased=False)
            cset = ax_slp2.contour(X_slp2, Y_slp2, Z_slp2, 7, zdir='z', offset=btn_limitz_slp2, linewidths=1,
                            cmap=cm.turbo, alpha=0.9)

            ax_slp2.set_zlim(btn_limitz_slp2, top_limitz_slp2)            
            ax_slp2.view_init(30, 240)
            ax_slp2.set_xlabel('X (mm)')
            ax_slp2.set_ylabel('Y (mm)')
            ax_slp2.set_zlabel(unit)

            ax_slp2.set_title('Abstracted tilting \n z = (%.2f) x + (%.2f) y + (%.2f), R$^2$ = %.2f' % (c[1], c[2],
                            c[0], R_sq1), fontsize=6)
            ax_slp2.text2D(0.2, 0.98, f'Tilt = %.7f°,' % angle_deg2, transform=ax_slp2.transAxes, fontsize=6)
            ax_slp2.text2D(0.53, 0.98, f'Skew = %.2f°' % non_acute, transform=ax_slp2.transAxes,
                            fontsize=6)
            fig_slp2.tight_layout(pad=1)
            
            plot_slp2 ="slope2"
            plt.savefig(plot_slp2, bbox_inches='tight')
            img_slp2 = Image.open("slope2.png")
            plot_slp2_resized = img_slp2.resize((356, 356))
            plot_slp2_resized.save("slope2_resized.png")     
            img_slp2 = ImageTk.PhotoImage(Image.open("slope2_resized.png"))
            self.canvas_graph.create_image(732, 380, anchor="nw", image=img_slp2)
            worksheet.insert_image('Q16', 'slope2.png', {'x_scale': 1.05, 'y_scale': 1.05})
            os.remove("slope2_resized.png")
            plt.clf()
            plt.close(fig_slp2)

# decompose_radius
            from scipy.optimize import curve_fit
            x_rad = df["A"]
            y_rad = df["B"]
            z_rad = df["Z"]
            x = df["radius (mm)"]
            y = z_rad - zz_slp1
            
            def objective(x, a, b, c, d):
                return a*x + b*x**2 + c*x**3 + d 
            popt, _ = curve_fit(objective, x, y)
            a, b, c, d = popt

            fig_rad = plt.figure(figsize=(2.5, 2.5))
            ax_rad = plt.axes(projection="3d")

            zz_rad = a*x + b*x**2 + c*x**3 + d
            Z_rad = griddata((x_rad, y_rad), zz_rad, (X_slp1, Y_slp1), method="cubic")

            my_fitting = np.polyfit(x, y, 3, full=True)
            coeff = my_fitting[0]
            SSE = my_fitting[1][0]
            diff = y - y.mean()
            square_diff = diff ** 2
            SST = square_diff.sum()
            R_sq2 = 1 - SSE/SST 

            btn_limitz_resi= btn_limitz3d
            top_limitz_resi = top_limitz3d

            Z_rad2 = Z_rad + np.mean(z3d) 
            ax_rad.contour3D(X_slp1, Y_slp1, Z_rad2, 450, cmap='turbo', alpha=1, antialiased=False)
            cset = ax_rad.contour(X_slp1, Y_slp1, Z_rad, 7, zdir='z', offset=btn_limitz_resi, linewidths=1,
                            cmap=cm.turbo, alpha=0.9)

            ax_rad.set_zlim(btn_limitz_resi, top_limitz_resi)                                                                                                                      
            ax_rad.view_init(30, 240)      
            ax_rad.set_xlabel('X (mm)')
            ax_rad.set_ylabel('Y (mm)')
            ax_rad.set_zlabel(unit)
            ax_rad.set_title(
                    'Abstracted radial pattern \n z = (%.2f) r + (%.4f) r$^2$ + (%.6f) r$^3$ + (%.2f), R$^2$ = %.2f'
                            % (a, b, c, d, R_sq2), fontsize=6)
                 
            fig_rad.tight_layout(pad=1)    
            plot_rad ="radius"
            plt.savefig(plot_rad, bbox_inches='tight')
            img_rad = Image.open("radius.png")

            plot_rad_resized = img_rad.resize((376, 356))
            plot_rad_resized.save("radius_resized.png")     
            img_rad = ImageTk.PhotoImage(Image.open("radius_resized.png"))  
            self.canvas_graph.create_image(1088, 380, anchor="nw", image=img_rad)
            worksheet.insert_image('V16', 'radius.png', {'x_scale': 1.05, 'y_scale': 1.05})
            os.remove("radius_resized.png")
            plt.clf()
            plt.close(fig_rad)

# decompose_residual
            fig_res = plt.figure(figsize=(2.5, 2.5))
            ax_res = plt.axes(projection="3d")
            
            degree_theta = float(self.rotation_entry.get())
            radian_rotation = math.radians(degree_theta)
            if (degree_theta == 0):
                x_res = df["A"]
                y_res = df["B"]
            else:
                x_orire = df["A"]
                y_orire = df["B"]
                x_res = math.cos(radian_rotation)*x_orire + math.sin(radian_rotation)*y_orire
                y_res = -math.sin(radian_rotation)*x_orire + math.cos(radian_rotation)*y_orire

            x_tmp = df[['x_tc', 'y_tc']].values.reshape(-1, 2)
            max_rows = max(len(x_slp1), len(y_slp1))
            x_ones = np.ones([max_rows, 1])
            x = np.append(x_ones, x_tmp, axis=1)
            y_tmp = z_slp1.tolist()
            y = np.array(y_tmp)

            def mse(coef, x, y):
                return np.mean((np.dot(x, coef) - y)**2)/2
            def gradients(coef, x, y):
                return np.mean(x.transpose()*(np.dot(x, coef) - y), axis=1)
            def multilinear_regression(coef, x, y, lr, b1=0.9, b2=0.999, epsilon=1e-8):
                prev_error = 0
                m_coef = np.zeros(coef.shape)
                v_coef = np.zeros(coef.shape)
                moment_m_coef = np.zeros(coef.shape)
                moment_v_coef = np.zeros(coef.shape)
                t = 0
                while True:
                    error = mse(coef, x, y)
                    if abs(error - prev_error) <= epsilon:
                        break
                    prev_error = error
                    grad = gradients(coef, x, y)
                    t += 1
                    m_coef = b1*m_coef + (1-b1)*grad
                    v_coef = b2*v_coef + (1-b2)*grad**2
                    moment_m_coef = m_coef / (1-b1**t)
                    moment_v_coef = v_coef / (1-b2**t)
                    delta = ((lr / moment_v_coef**0.5 + 1e-8)*(b1*moment_m_coef + (1-b1)*grad/(1-b1**t)))
                    coef = np.subtract(coef, delta)
                return coef
            
            coef = np.array([0, 0, 0])
            c = multilinear_regression(coef, x, y, 1e-1)
            
            x_res_grid = np.linspace(np.min(x_res), np.max(x_res), 120)
            y_res_grid = np.linspace(np.min(y_res), np.max(y_res), 120)
            X_res, Y_res = np.meshgrid(x_res_grid, y_res_grid)
            
            z_res = df["Z"]
            zz_res = c[1]*x_res + c[2]*y_res + c[0]
            Z_res = griddata((x_res, y_res), z_res, (X_res, Y_res), method="cubic")
            ZZ_res = griddata((x_res, y_res), zz_res, (X_res, Y_res), method="linear")

            check_box = self.var_limits.get() 
            if (check_box == 1):
                limits_z_pre = self.limits_entry.get()
                limits_z = abs(float(limits_z_pre))
                btn_limitz_res = np.mean(z_res) - limits_z/2
                top_limitz_res = np.mean(z_res) + limits_z/2       
            else:
                top_limitz_res = np.mean(z_res) + (max(z_res)-min(z_res))*1.5
                btn_limitz_res = np.mean(z_res) - (max(z_res)-min(z_res))*4.5
         
            ax_res.contour3D(X_res, Y_res, Z_res - ZZ_res + c[0] - Z_rad, 450, cmap='turbo', alpha=1,
                            antialiased=False)
            cset = ax_res.contour(X_res, Y_res, Z_res - ZZ_res + c[0] - Z_rad, 7, zdir='z', offset=btn_limitz_res,
                            linewidths=1, cmap=cm.turbo, alpha=1)

            z_after = z_res - zz_res + c[0] - zz_rad
            ave_res = round(np.mean(z_after), 1)
            max_res = round(max(z_after))
            min_res = round(min(z_after))
            nonu_res = round(0.5*100*(max_res - min_res)/abs(ave_res), 2)
        
            ax_res.view_init(30, 240)
            ax_res.set_zlim(btn_limitz_res, top_limitz_res)
            ax_res.set_xlabel('X (mm)')
            ax_res.set_ylabel('Y (mm)')
            ax_res.set_zlabel(unit)
                 
            ax_res.set_title(
                            'After both tilting and radial pattern are corrected \n Mean = %.1f Å, NonU = %.2f pct'
                            %(ave_res, nonu_res), fontsize=6)
            fig_res.tight_layout(pad=1)    
            plot_res ="residual"
            plt.savefig(plot_res, bbox_inches='tight')
            img_res = Image.open("residual.png")
            plot_res_resized = img_res.resize((356, 356))
            plot_res_resized.save("residual_resized.png")     
            img_res = ImageTk.PhotoImage(Image.open("residual_resized.png"))
            self.canvas_graph.create_image(1088, 0, anchor="nw", image=img_res)
            worksheet.insert_image('V1', 'residual.png', {'x_scale': 1.05, 'y_scale': 1.05})
            os.remove("residual_resized.png")
            plt.clf()
            plt.close(fig_res)
        else:
            pass
        workbook.close()

        if (check_box2 == 1):
            os.remove("residual.png")
            os.remove("radius.png")
            os.remove("slope2.png")
            os.remove("slope1.png")
            os.remove("cdf.png")
            os.remove("hist.png")
            os.remove("scatter2.png")
            os.remove("scatter1.png")
            os.remove("3d.png")
            os.remove("cross_section2.png")
            os.remove("cross_section1.png")
            os.remove("contour.png")
        else:
            os.remove("3d.png")
            os.remove("cross_section2.png")
            os.remove("cross_section1.png")
            os.remove("contour.png")
        
        tk.mainloop()

    def copy_clipboard(self):
        fig1 = plt.figure(figsize=(2.5, 2.5))
        spec1 = gridspec.GridSpec(ncols=1, nrows=1)
        ax1 = fig1.add_subplot(spec1[0, 0])
        
        degree_theta = float(self.rotation_entry.get())
        radian_theta = math.radians(degree_theta)
        if (degree_theta == 0):
            x2d = df["A"]
            y2d = df["B"]
        else:
            x_original = df["A"]
            y_original = df["B"]        
            x2d = math.cos(radian_theta)*x_original + math.sin(radian_theta)*y_original
            y2d = - math.sin(radian_theta)*x_original + math.cos(radian_theta)*y_original
        z2d = df["Z"]

        x_grid = np.linspace(np.min(x2d), np.max(x2d), 300)
        y_grid = np.linspace(np.min(y2d), np.max(y2d), 300)
        X2d, Y2d = np.meshgrid(x_grid, y_grid)
        Z2d = griddata((x2d, y2d), z2d, (X2d, Y2d), method="cubic")
        
        Ave = round(np.mean(z2d), 1)
        Std = np.std(z2d)
        Std_percent = round(100*Std/abs(Ave), 2)
        Max = round(max(z2d), 1)
        Min = round(min(z2d), 1)
        Nonu = round(0.5*100*(Max-Min)/abs(Ave), 2)
        Range = round(max(z2d) - min(z2d), 1)

        value1 = self.variable_cal.get()
        if (value1 == '(A-B)/t'):
            time_pre = self.entry_run_time.get()
            time = float(time_pre)
            Rate_pre = abs(Ave)*(60/time)
            Rate = round(Rate_pre, 1)         
            plt.text(-75,-330, 'Rate (' + unit + '/min)')
            plt.text(45,-330, Rate)
        elif (value1 == '(B-A)/t'):
            time_pre = self.entry_run_time.get()
            time = float(time_pre)
            Rate_pre = abs(Ave)*(60/time)
            Rate = round(Rate_pre, 1)         
            plt.text(-75,-330, 'Rate (' + unit + '/min)')
            plt.text(45,-330, Rate)
        else:
            pass
        
        plt.text(-75,-240, 'Mean ' + unit)
        plt.text(-75,-255, 'NonU %')
        plt.text(-75,-270, 'StdD %')
        plt.text(-75,-285, 'Max')
        plt.text(-75,-300, 'Min')
        plt.text(-75,-315, 'Range')
        plt.text(45,-240, Ave)
        plt.text(45,-255, Nonu)                                      
        plt.text(45,-270, Std_percent)
        plt.text(45,-285, Max)
        plt.text(45,-300, Min)
        plt.text(45,-315, Range)

        check_box = self.var_limits.get() 
        if (check_box == 1):
            limits_bt_pre = self.limits_entry.get()
            limits_bt = abs(float(limits_bt_pre))
            limits_b = Ave - limits_bt/2
            limits_t = Ave + limits_bt/2
            num = int(self.entry_contour.get())
            if (num >100):
                num = 100
                levels = np.linspace(limits_b, limits_t, num)
            else:
                levels = np.linspace(limits_b, limits_t, num)
            cp = plt.contourf(X2d, Y2d, Z2d, levels = levels, cmap=plt.cm.turbo, alpha=0.95)
            fig1.colorbar(cp, ax=ax1, shrink=0.5, orientation="vertical")
        else:
            contour = int(self.entry_contour.get()) 
            if (contour >100):
                contour = 100
            else:
                pass
            cp = plt.contourf(X2d, Y2d, Z2d, contour, cmap=plt.cm.turbo, alpha=0.95)
            cbar1 = fig1.colorbar(cp, ax=ax1, shrink=0.5, orientation="vertical")
            cbar1.ax.locator_params(nbins=6)
            
        value = self.radio_sigma.get()
        if (value == 1):
            plt.text(-75,-345, 'Points removed')
            plt.text(45,-345, N1)
        else:
            plt.text(-75,-345, 'Sigma:')
            plt.text(45,-345, sigma2)
        
        circ = Circle((0, 0), 150, facecolor='None', edgecolor='black', lw=0.2, alpha=0.4)
        ax1.add_patch(circ)
        ax1.set_aspect('equal', adjustable='box')

        var3 = self.variable3.get()
        if (var3 == "Value"):
            for x2d, y2d, z2d in zip(x2d, y2d, z2d):
                label = "{:.0f}".format(z2d)
                plt.annotate(label, (x2d, y2d), textcoords="offset points", xytext=(0, -3), ha='center', 
                             fontsize=4.3, alpha=0.85)
                ax1.scatter(x2d, y2d, marker='o', s=1, color='k', alpha=0.0001)
        elif (var3 == "Dot"):
            for x2d, y2d, z2d in zip(x2d, y2d, z2d):
                ax1.scatter(x2d, y2d, marker='o', s=1, color='k', alpha=0.3)
        elif (var3 == "None"):
            for x2d, y2d, z2d in zip(x2d, y2d, z2d):
                label = "{:.0f}".format(z2d)
                ax1.scatter(x2d, y2d, marker='o', s=1, color='k', alpha=0.0001)
        else:
            for x2d, y2d, z2d in zip(x2d, y2d, z2d):
                label = "{:.0f}".format(z2d)
                if (z2d > Ave):
                    ax1.scatter(x2d, y2d, marker='+', s=13, linewidths=0.3, color='k', alpha=0.7)
                else:
                    ax1.scatter(x2d, y2d, marker='_', s=13, linewidths=0.3, color='k', alpha=0.7)

        ax1.set_xlabel('X (mm)')
        ax1.set_ylabel('Y (mm)')    
        ax1.set_title(id + '    ' + film1)
    
        fig1.tight_layout(pad=0.1)
        plot_id11 = "contour1"
        plt.savefig(plot_id11, bbox_inches='tight')        
        
        def send_to_clipboard1(clip_type, data):
            win32clipboard.OpenClipboard()
            win32clipboard.EmptyClipboard()
            win32clipboard.SetClipboardData(clip_type, data)
            win32clipboard.CloseClipboard()

        image = Image.open("contour1.png")
        output = BytesIO()
        image.convert("RGB").save(output, "BMP")
        data = output.getvalue()[14:]
        output.close()
        send_to_clipboard1(win32clipboard.CF_DIB, data)
        os.remove("contour1.png")

        tk.mainloop()
                
def main():
    my_gui = app_gui()
main()       

# li.hou2009@gmail.com
# in memory of dear father Chongjian Hou (1928-2021) 
# https://lihou0427.github.io/solidplasma/
